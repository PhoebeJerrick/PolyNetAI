#!/usr/bin/env python3
"""
Polymarket 交易明细 Excel 持仓分析 + 盈亏汇总。

插入到"成交价格"后的列（若输入中存在）：
    同向成交均价 — 表示本方向（Up/Down）到当前行为止所有成交的份数加权均价（含买入和卖出），首笔即为当前成交价；
    同向成交价波动幅度(%) — 表示相对本周期内上一次同结果代币成交价的涨跌幅百分比，首笔为空；
    相对于加权均价的价格波动百分比 — 表示当前行成交价相对该方向当前加权均价的涨跌幅百分比；
    其后为 Up积累份数 / Up持仓成本 / Up的加权均价 / Down积累份数 / Down持仓成本 / Down加权均价 /
    当前总持仓份数 / 持仓价值占比 / 持仓价值加/减仓百分比 / 净持仓份数 / 净持仓成本差额 / 净持仓方向 / 持仓异常 /
    Up已成交差价盈亏 / Down已成交差价盈亏 / 浮动盈亏

净持仓成本差额 = Up持仓成本 − Down持仓成本。
若出现卖出超过当前已持有份数，不再把超卖部分继续计入负持仓；
该超卖部分会被单独标记到「持仓异常」列，并且不参与持仓/持仓价值/浮动盈亏计算。

持仓价值加/减仓百分比 = 本笔成交对应的持仓成本变动额 / 变动前该方向持仓成本。
加仓记为正，减仓记为负；若该方向变动前没有持仓，则留空。

输出前会删除列：市场ID、交易类型、订单类型、备注、时间；
用「下注时间距开盘差(分，秒)」代替「时间」并置于首列。

差价盈亏 = 每笔卖出时 shares × (卖出价 − 该方向加权平均买入价)，
仅在该方向有卖出的行才显示数值，其余行留空。

每个时间周期末尾的小计行额外包含：
  未平仓UP盈亏 / 未平仓Down盈亏 / 周期净利润

结算判定：周期内最后成交价 > 0.5 的方向胜出，
赢家按 1.0 结算、输家按 0.0 结算。
"""

from __future__ import annotations

import argparse
import math
import sys
from pathlib import Path
from typing import Iterable, Sequence

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import yaml


# ── column-name aliases ──────────────────────────────────────────────
PRICE_ALIASES = ("成交价格", "价格", "price", "avgprice")
QTY_ALIASES = ("投注份数", "份数", "成交份数", "成交数量", "数量", "shares", "size", "qty", "amount")
OUTCOME_ALIASES = ("结果代币类型", "方向", "结果", "币种方向", "Outcome", "outcome", "token", "token_side")
ACTION_ALIASES = ("操作方向", "操作", "交易类型", "买卖", "买卖方向", "action", "side", "trade_side")
COIN_ALIASES = ("币种", "标的", "资产", "代码", "symbol", "ticker", "coin")
CYCLE_ALIASES = (
    "时间周期", "市场周期", "周期", "结算周期", "轮次", "场次",
    "market_cycle", "cycle", "event_start_time", "开始时间",
    "市场标题", "市场名称", "market",
)

# 代替「时间」列；用于小计行标记与冻结列参考
BET_TIME_ALIASES = (
    "下注时间距开盘差(分，秒)",
    "下注时间距开盘差（分，秒）",
    "下注时间距开盘差",
    "距开盘差",
)

# 输出时删除的列
DROP_OUTPUT_COLUMNS = ("市场ID", "交易类型", "订单类型", "备注", "时间")

BUY_KEYWORDS = ("buy", "买", "买入", "开仓", "加仓", "bought", "open")
SELL_KEYWORDS = ("sell", "卖", "卖出", "平仓", "减仓", "close", "closed", "止盈", "止损")
UP_KEYWORDS = ("up", "yes", "涨", "看涨", "做多", "多")
DOWN_KEYWORDS = ("down", "no", "跌", "看跌", "做空", "空")

# ── column groups ────────────────────────────────────────────────────
# 可选；由批量回放等上游提供，insert_columns 会紧挨「成交价格」插入，再插入 INSERTED_COLUMNS
EXTRA_COLUMNS_AFTER_PRICE = ("同向成交均价", "同向成交价波动幅度(%)", "相对于加权均价的价格波动百分比")

INSERTED_COLUMNS = (
    "Up积累份数",
    "Up持仓成本",
    "Up的加权均价",
    "Down积累份数",
    "Down持仓成本",
    "Down加权均价",
    "当前总持仓份数",
    "持仓价值占比",
    "持仓价值加/减仓百分比",
    "净持仓份数",
    "净持仓成本差额",
    "净持仓方向",
    "持仓异常",
    "Up已成交差价盈亏",
    "Down已成交差价盈亏",
    "浮动盈亏",
)
SETTLEMENT_COLUMNS = ("未平仓UP盈亏", "未平仓Down盈亏", "周期净利润")
SETTLEMENT_META_COLUMNS = ("最终Winner方向",)
ALL_PNL_COLUMNS = ("Up已成交差价盈亏", "Down已成交差价盈亏", "浮动盈亏") + SETTLEMENT_COLUMNS
DEFAULT_POSITION_VALUE_DENOMINATOR = 85.0

# ── styles ───────────────────────────────────────────────────────────
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)

UP_FILL = PatternFill(fill_type="solid", fgColor="E2F0D9")
DOWN_FILL = PatternFill(fill_type="solid", fgColor="FCE4D6")
TOTAL_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
NET_POS_FILL = PatternFill(fill_type="solid", fgColor="E8DAEF")
DIRECTION_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
ANOMALY_FILL = PatternFill(fill_type="solid", fgColor="FDE9D9")
UP_PNL_FILL = PatternFill(fill_type="solid", fgColor="E2EFDA")
DN_PNL_FILL = PatternFill(fill_type="solid", fgColor="FBE5D6")

NET_DIRECTION_FILLS = {
    "Up": PatternFill(fill_type="solid", fgColor="C6E0B4"),
    "Down": PatternFill(fill_type="solid", fgColor="F4CCCC"),
    "平衡": PatternFill(fill_type="solid", fgColor="FFE699"),
    "空仓": PatternFill(fill_type="solid", fgColor="D9D9D9"),
}

SUBTOTAL_FILL = PatternFill(fill_type="solid", fgColor="D6E4F0")
SUBTOTAL_FONT = Font(bold=True)
SUBTOTAL_BORDER = Border(
    top=Side(style="thin", color="4472C4"),
    bottom=Side(style="thin", color="4472C4"),
)

PNL_HEADER_FILL = PatternFill(fill_type="solid", fgColor="4472C4")
PNL_HEADER_FONT = Font(color="FFFFFF", bold=True)
PNL_POS_FONT = Font(bold=True, color="006100")
PNL_NEG_FONT = Font(bold=True, color="9C0006")

CYCLE_FILLS = (
    PatternFill(fill_type="solid", fgColor="F2F7FB"),
    None,
)

COL_FILLS = {
    "Up积累份数": UP_FILL,
    "Up持仓成本": UP_FILL,
    "Up的加权均价": UP_FILL,
    "Down积累份数": DOWN_FILL,
    "Down持仓成本": DOWN_FILL,
    "Down加权均价": DOWN_FILL,
    "当前总持仓份数": TOTAL_FILL,
    "持仓价值占比": TOTAL_FILL,
    "持仓价值加/减仓百分比": TOTAL_FILL,
    "净持仓份数": NET_POS_FILL,
    "净持仓成本差额": NET_POS_FILL,
    "净持仓方向": DIRECTION_FILL,
    "持仓异常": ANOMALY_FILL,
    "Up已成交差价盈亏": UP_PNL_FILL,
    "Down已成交差价盈亏": DN_PNL_FILL,
    "浮动盈亏": PatternFill(fill_type="solid", fgColor="E7E9FF"),
}


# ═══════════════════════ Utility helpers ═════════════════════════════

def normalize_text(value: object) -> str:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return ""
    text = str(value).strip().lower()
    for ch in (" ", "_", "-", "/", "\\", "（", "）", "(", ")", "[", "]"):
        text = text.replace(ch, "")
    return text


def split_group_cols(raw: str | None) -> list[str] | None:
    if not raw:
        return None
    return [p.strip() for p in raw.replace("，", ",").split(",") if p.strip()]


def find_column(
    columns: Sequence[str], aliases: Iterable[str],
    explicit: str | None = None, required: bool = True,
) -> str | None:
    if explicit:
        if explicit not in columns:
            raise ValueError(f"指定列不存在: {explicit}")
        return explicit
    nm = {normalize_text(c): c for c in columns}
    for a in aliases:
        na = normalize_text(a)
        if na in nm:
            return nm[na]
    for a in aliases:
        na = normalize_text(a)
        for cand in nm:
            if na and na in cand:
                return nm[cand]
    if required:
        raise ValueError("未找到列。候选: " + ", ".join(str(a) for a in aliases)
                         + "\n表头: " + ", ".join(str(c) for c in columns))
    return None


def parse_num(val: object, rn: int, cn: str) -> float:
    if val is None or (isinstance(val, float) and math.isnan(val)):
        raise ValueError(f"第 {rn} 行 `{cn}` 为空。")
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return float(val)
    try:
        return float(str(val).strip().replace(",", ""))
    except ValueError as e:
        raise ValueError(f"第 {rn} 行 `{cn}` 无法解析: {val!r}") from e


def parse_outcome(val: object, rn: int) -> str:
    t = normalize_text(val)
    if not t:
        raise ValueError(f"第 {rn} 行无方向。")
    if any(k in t for k in UP_KEYWORDS):
        return "up"
    if any(k in t for k in DOWN_KEYWORDS):
        return "down"
    raise ValueError(f"第 {rn} 行无法识别方向: {val!r}")


def parse_action(val: object, rn: int) -> int | None:
    t = normalize_text(val)
    if not t:
        return None
    if any(k in t for k in BUY_KEYWORDS):
        return 1
    if any(k in t for k in SELL_KEYWORDS):
        return -1
    raise ValueError(f"第 {rn} 行无法识别买卖: {val!r}")


def f3(v: float) -> int | float:
    r = round(v, 3)
    if abs(r) < 1e-10:
        return 0
    if abs(r - round(r)) < 1e-10:
        return int(round(r))
    return r


def net_dir(up: float, dn: float) -> str:
    ru, rd = round(up, 10), round(dn, 10)
    if abs(ru + rd) < 1e-10:
        return "空仓"
    if abs(ru - rd) < 1e-10:
        return "平衡"
    return "Up" if ru > rd else "Down"


def round_all(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for c in out.columns:
        out[c] = out[c].apply(
            lambda v: round(float(v), 3)
            if isinstance(v, (int, float)) and not isinstance(v, bool) and pd.notna(v)
            else v
        )
    return out


def preprocess_dataframe(df: pd.DataFrame) -> tuple[pd.DataFrame, str]:
    """删除指定列；用「下注时间距开盘差」代替「时间」并置于首列。"""
    out = df.copy()
    for col in DROP_OUTPUT_COLUMNS:
        if col in out.columns:
            out = out.drop(columns=[col])
    bet_col = find_column(list(out.columns), BET_TIME_ALIASES, required=False)
    if bet_col:
        order = [bet_col] + [c for c in out.columns if c != bet_col]
        out = out[order]
        marker = bet_col
    else:
        marker = str(out.columns[0])
    return out, marker


def net_position_cost_gap(up_bal: float, dn_bal: float, up_avg: float, dn_avg: float) -> int | float:
    """成本口径净敞口：up_cost − down_cost。"""
    return f3(up_bal * up_avg - dn_bal * dn_avg)


def same_outcome_price_move_pct(current_price: float, previous_price: float) -> float | str:
    if previous_price <= 1e-12:
        return ""
    return current_price / previous_price - 1.0


def relative_to_weighted_avg_price_move_pct(current_price: float, weighted_avg_price: float) -> float | str:
    if weighted_avg_price <= 1e-12:
        return ""
    return current_price / weighted_avg_price - 1.0


def position_value_change_pct(
    *,
    signed_qty: float,
    shares: float,
    price: float,
    pre_held: float,
    pre_cost: float,
    pre_avg: float,
) -> float | str:
    if pre_cost <= 1e-12:
        return ""
    if signed_qty > 0:
        return (shares * price) / pre_cost
    matched = min(shares, pre_held)
    if matched <= 1e-12:
        return ""
    return -((matched * pre_avg) / pre_cost)


# ═══════════════════════ Column resolution ═══════════════════════════

def choose_groups(df: pd.DataFrame, args: argparse.Namespace) -> list[str]:
    if args.group_cols:
        miss = [c for c in args.group_cols if c not in df.columns]
        if miss:
            raise ValueError(f"分组列不存在: {', '.join(miss)}")
        return args.group_cols
    coin = find_column(df.columns, COIN_ALIASES, args.coin_col, required=False)
    cyc = find_column(df.columns, CYCLE_ALIASES, args.cycle_col, required=True)
    return [coin, cyc] if coin else [cyc]


def resolve_cols(df: pd.DataFrame, args: argparse.Namespace) -> dict[str, str | None]:
    return {
        "qty": find_column(df.columns, QTY_ALIASES, args.qty_col, required=True),
        "outcome": find_column(df.columns, OUTCOME_ALIASES, args.outcome_col, required=True),
        "action": find_column(df.columns, ACTION_ALIASES, args.action_col, required=False),
        "price": find_column(df.columns, PRICE_ALIASES, args.price_col, required=True),
    }


# ═══════════════════════ P&L helpers ═════════════════════════════════

def new_bucket() -> dict[str, float]:
    return {
        "up_held": 0.0, "up_cost": 0.0, "up_avg": 0.0,
        "up_realized": 0.0,
        "up_fill_cost": 0.0, "up_fill_shares": 0.0,
        "dn_held": 0.0, "dn_cost": 0.0, "dn_avg": 0.0,
        "dn_realized": 0.0,
        "dn_fill_cost": 0.0, "dn_fill_shares": 0.0,
        "anomaly_count": 0.0,
        "last_up": 0.0, "last_dn": 0.0,
    }


def format_position_anomaly(direction: str, excess_shares: float) -> str:
    return f"{direction}卖出超持仓 {f3(excess_shares)}份"


def winner(s: dict[str, float]) -> str | None:
    """基于用户最后成交价推断赢家（仅作 fallback）。"""
    u, d = s["last_up"], s["last_dn"]
    if u > 0 and d > 0:
        return "up" if u >= d else "down"
    if u > 0:
        return "up" if u > 0.5 else "down"
    if d > 0:
        return "down" if d > 0.5 else "up"
    return None


def settlement_summary(
    s: dict[str, float],
    override_winner: str | None = None,
) -> dict[str, float]:
    rem_up = s["up_held"]
    rem_dn = s["dn_held"]
    w = override_winner or winner(s)

    up_stl = dn_stl = 0.0
    if w == "up":
        if rem_up > 1e-10:
            up_stl = rem_up * (1.0 - s["up_avg"])
        if rem_dn > 1e-10:
            dn_stl = -(rem_dn * s["dn_avg"])
    elif w == "down":
        if rem_up > 1e-10:
            up_stl = -(rem_up * s["up_avg"])
        if rem_dn > 1e-10:
            dn_stl = rem_dn * (1.0 - s["dn_avg"])

    return {
        "未平仓UP盈亏": f3(up_stl),
        "未平仓Down盈亏": f3(dn_stl),
        "up_realized_total": f3(s["up_realized"]),
        "dn_realized_total": f3(s["dn_realized"]),
        "周期净利润": f3(s["up_realized"] + s["dn_realized"] + up_stl + dn_stl),
        "最终Winner方向": "Up" if w == "up" else ("Down" if w == "down" else ""),
    }


# ═══════════════════════ Data processing ═════════════════════════════

def insert_columns(df: pd.DataFrame, price_col: str) -> pd.DataFrame:
    out = df.copy()
    saved: dict[str, pd.Series] = {}
    for c in INSERTED_COLUMNS:
        if c in out.columns:
            saved[c] = out.pop(c)
        else:
            raise ValueError(f"缺少 `{c}` 列。")
    extra_saved: dict[str, pd.Series] = {}
    for c in EXTRA_COLUMNS_AFTER_PRICE:
        if c in out.columns:
            extra_saved[c] = out.pop(c)
    pos = int(out.columns.get_loc(price_col)) + 1
    offset = 0
    for c, series in extra_saved.items():
        out.insert(pos + offset, c, series)
        offset += 1
    for i, c in enumerate(INSERTED_COLUMNS):
        out.insert(pos + offset + i, c, saved[c])
    return out


def append_subtotals(
    df: pd.DataFrame,
    cycle_col: str,
    pnl_stats: dict[tuple[object, ...], dict[str, float]],
    marker_col: str,
    group_cols: list[str],
    winner_override: str | None = None,
) -> pd.DataFrame:
    if df.empty:
        return df

    all_cols = list(df.columns) + [c for c in SETTLEMENT_COLUMNS + SETTLEMENT_META_COLUMNS if c not in df.columns]
    rows: list[dict[str, object]] = []

    for cyc_val, grp in df.groupby(cycle_col, sort=False):
        rows.extend(grp.to_dict(orient="records"))

        last = grp.iloc[-1]
        sub: dict[str, object] = {c: "" for c in all_cols}
        sub[marker_col] = "【周期小计】"
        sub[cycle_col] = cyc_val

        for c in INSERTED_COLUMNS:
            sub[c] = last.get(c, "")

        # 与 compute 中 pnl_s 的键一致：tuple(row[g] for g in groups)
        key = tuple(last[c] for c in group_cols)
        stats = pnl_stats.get(key)
        if stats:
            sm = settlement_summary(stats, override_winner=winner_override)
            sub["Up已成交差价盈亏"] = sm["up_realized_total"]
            sub["Down已成交差价盈亏"] = sm["dn_realized_total"]
            sub["未平仓UP盈亏"] = sm["未平仓UP盈亏"]
            sub["未平仓Down盈亏"] = sm["未平仓Down盈亏"]
            sub["周期净利润"] = sm["周期净利润"]
            sub["最终Winner方向"] = sm["最终Winner方向"]
            anomaly_count = int(round(float(stats.get("anomaly_count", 0.0))))
            if anomaly_count > 0:
                sub["持仓异常"] = f"本周期持仓异常 {anomaly_count} 笔"

        rows.append(sub)

    return pd.DataFrame(rows, columns=all_cols)


def compute(
    df: pd.DataFrame,
    args: argparse.Namespace,
    winner_override: str | None = None,
) -> tuple[pd.DataFrame, dict[str, object]]:
    if df.empty:
        raise ValueError("工作表为空。")

    df, marker_col = preprocess_dataframe(df)

    groups = choose_groups(df, args)
    rc = resolve_cols(df, args)
    qty_c = str(rc["qty"])
    out_c = str(rc["outcome"])
    act_c = rc["action"]
    prc_c = str(rc["price"])
    cyc_c = next(
        (c for c in groups if normalize_text(c) == normalize_text(args.cycle_col or "时间周期")),
        groups[-1],
    )

    pnl_s: dict[tuple, dict[str, float]] = {}

    up_v: list[object] = []
    up_cost_v: list[object] = []
    up_avg_v: list[object] = []
    dn_v: list[object] = []
    dn_cost_v: list[object] = []
    dn_avg_v: list[object] = []
    tot_v: list[object] = []
    pos_ratio_v: list[object] = []
    pos_value_delta_pct_v: list[object] = []
    net_v: list[object] = []
    net_val_v: list[object] = []
    dir_v: list[str] = []
    anomaly_v: list[str] = []
    same_outcome_move_v: list[object] = []
    weighted_avg_move_v: list[object] = []
    same_dir_avg_v: list[object] = []
    up_pnl_v: list[object] = []
    dn_pnl_v: list[object] = []
    float_pnl_v: list[object] = []

    for i, row in df.iterrows():
        rn = i + 2
        gk = tuple(row[c] for c in groups)
        if any(pd.isna(x) for x in gk):
            raise ValueError(f"第 {rn} 行分组字段为空。")

        qty = parse_num(row[qty_c], rn, qty_c)
        outcome = parse_outcome(row[out_c], rn)
        act = parse_action(row[act_c], rn) if act_c else None

        if act is None:
            signed = qty if qty < 0 else (_ for _ in ()).throw(
                ValueError(f"第 {rn} 行缺买卖列且份数非负。"))
        else:
            signed = abs(qty) * act

        price = float(row[prc_c]) if pd.notna(row[prc_c]) else 0.0
        b = pnl_s.setdefault(gk, new_bucket())
        shares = abs(signed)
        pre_up_held = b["up_held"]
        pre_up_cost = b["up_cost"]
        pre_up_avg = b["up_avg"]
        pre_dn_held = b["dn_held"]
        pre_dn_cost = b["dn_cost"]
        pre_dn_avg = b["dn_avg"]
        pre_last_up = b["last_up"]
        pre_last_dn = b["last_dn"]
        trade_pnl_up = ""
        trade_pnl_dn = ""
        row_anomalies: list[str] = []
        same_outcome_move = ""
        position_value_delta_pct = ""

        if outcome == "up":
            same_outcome_move = same_outcome_price_move_pct(price, pre_last_up)
            position_value_delta_pct = position_value_change_pct(
                signed_qty=signed,
                shares=shares,
                price=price,
                pre_held=pre_up_held,
                pre_cost=pre_up_cost,
                pre_avg=pre_up_avg,
            )
            if signed > 0:
                b["up_held"] += shares
                b["up_cost"] += shares * price
                b["up_avg"] = b["up_cost"] / b["up_held"] if b["up_held"] > 1e-10 else 0.0
            else:
                matched = min(shares, b["up_held"])
                excess = shares - matched
                if matched > 1e-10:
                    tp = matched * (price - b["up_avg"])
                    b["up_realized"] += tp
                    trade_pnl_up = f3(tp)
                    b["up_held"] -= matched
                    if b["up_held"] > 1e-10:
                        b["up_cost"] = b["up_held"] * b["up_avg"]
                    else:
                        b["up_held"] = 0.0
                        b["up_cost"] = 0.0
                        b["up_avg"] = 0.0
                if excess > 1e-10:
                    b["anomaly_count"] += 1.0
                    row_anomalies.append(format_position_anomaly("Up", excess))
            b["up_fill_cost"] += shares * price
            b["up_fill_shares"] += shares
            b["last_up"] = price
        else:
            same_outcome_move = same_outcome_price_move_pct(price, pre_last_dn)
            position_value_delta_pct = position_value_change_pct(
                signed_qty=signed,
                shares=shares,
                price=price,
                pre_held=pre_dn_held,
                pre_cost=pre_dn_cost,
                pre_avg=pre_dn_avg,
            )
            if signed > 0:
                b["dn_held"] += shares
                b["dn_cost"] += shares * price
                b["dn_avg"] = b["dn_cost"] / b["dn_held"] if b["dn_held"] > 1e-10 else 0.0
            else:
                matched = min(shares, b["dn_held"])
                excess = shares - matched
                if matched > 1e-10:
                    tp = matched * (price - b["dn_avg"])
                    b["dn_realized"] += tp
                    trade_pnl_dn = f3(tp)
                    b["dn_held"] -= matched
                    if b["dn_held"] > 1e-10:
                        b["dn_cost"] = b["dn_held"] * b["dn_avg"]
                    else:
                        b["dn_held"] = 0.0
                        b["dn_cost"] = 0.0
                        b["dn_avg"] = 0.0
                if excess > 1e-10:
                    b["anomaly_count"] += 1.0
                    row_anomalies.append(format_position_anomaly("Down", excess))
            b["dn_fill_cost"] += shares * price
            b["dn_fill_shares"] += shares
            b["last_dn"] = price

        up_held = b["up_held"]
        dn_held = b["dn_held"]

        if outcome == "up":
            same_dir_avg = f3(b["up_fill_cost"] / b["up_fill_shares"]) if b["up_fill_shares"] > 1e-10 else ""
        else:
            same_dir_avg = f3(b["dn_fill_cost"] / b["dn_fill_shares"]) if b["dn_fill_shares"] > 1e-10 else ""
        same_dir_avg_v.append(same_dir_avg)

        up_v.append(f3(up_held))
        up_cost_v.append(f3(b["up_cost"]))
        dn_v.append(f3(dn_held))
        dn_cost_v.append(f3(b["dn_cost"]))
        tot_v.append(f3(up_held + dn_held))
        net_v.append(f3(up_held - dn_held))
        dir_v.append(net_dir(up_held, dn_held))
        anomaly_v.append("；".join(row_anomalies))
        same_outcome_move_v.append(same_outcome_move)
        weighted_avg_move_v.append(
            relative_to_weighted_avg_price_move_pct(
                price,
                b["up_avg"] if outcome == "up" else b["dn_avg"],
            )
        )
        pos_value_delta_pct_v.append(position_value_delta_pct)

        # 持仓价值占比 = 当前总持仓成本价值 / 分母（默认对齐 position.max_position_value）
        ratio_denominator = float(
            getattr(args, "position_value_denominator", DEFAULT_POSITION_VALUE_DENOMINATOR)
            or DEFAULT_POSITION_VALUE_DENOMINATOR
        )
        if ratio_denominator <= 1e-12:
            ratio_denominator = DEFAULT_POSITION_VALUE_DENOMINATOR
        total_cost_value = b["up_cost"] + b["dn_cost"]
        pos_ratio_v.append(total_cost_value / ratio_denominator)

        up_pnl_v.append(trade_pnl_up)
        dn_pnl_v.append(trade_pnl_dn)

        # 浮动盈亏：基于“当前行的成交价格”推导另一侧价格（二元合约：Up + Down = 1），
        # 并对当前已持仓部分做市价估值 (held * (px - avg_cost))。
        if outcome == "up":
            up_px = price
            dn_px = 1.0 - price
        else:
            dn_px = price
            up_px = 1.0 - price
        float_pnl = b["up_held"] * (up_px - b["up_avg"]) + b["dn_held"] * (dn_px - b["dn_avg"])
        float_pnl_v.append(f3(float_pnl))

        up_avg_v.append(f3(b["up_avg"]) if b["up_held"] > 1e-10 else 0)
        dn_avg_v.append(f3(b["dn_avg"]) if b["dn_held"] > 1e-10 else 0)
        net_val_v.append(net_position_cost_gap(up_held, dn_held, b["up_avg"], b["dn_avg"]))

    res = df.copy()
    res["同向成交均价"] = same_dir_avg_v
    res["同向成交价波动幅度(%)"] = same_outcome_move_v
    res["相对于加权均价的价格波动百分比"] = weighted_avg_move_v
    res["Up积累份数"] = up_v
    res["Up持仓成本"] = up_cost_v
    res["Up的加权均价"] = up_avg_v
    res["Down积累份数"] = dn_v
    res["Down持仓成本"] = dn_cost_v
    res["Down加权均价"] = dn_avg_v
    res["当前总持仓份数"] = tot_v
    res["持仓价值占比"] = pos_ratio_v
    res["持仓价值加/减仓百分比"] = pos_value_delta_pct_v
    res["净持仓份数"] = net_v
    res["净持仓成本差额"] = net_val_v
    res["净持仓方向"] = dir_v
    res["持仓异常"] = anomaly_v
    res["Up已成交差价盈亏"] = up_pnl_v
    res["Down已成交差价盈亏"] = dn_pnl_v
    res["浮动盈亏"] = float_pnl_v
    res = insert_columns(res, prc_c)
    res = append_subtotals(res, cyc_c, pnl_s, marker_col, groups, winner_override=winner_override)
    res = round_all(res)

    return res, {
        "group_cols": groups,
        "cycle_col": cyc_c,
        "qty_col": qty_c,
        "outcome_col": out_c,
        "action_col": act_c,
        "price_col": prc_c,
        "marker_col": marker_col,
    }


# ═══════════════════════ Excel formatting ════════════════════════════

def format_ws(ws, marker_col: str | None = None) -> None:
    ws.freeze_panes = "A2"

    cm: dict[str, int] = {}
    for idx, cell in enumerate(ws[1], 1):
        cm[cell.value] = idx

    for cell in ws[1]:
        v = cell.value
        if v in INSERTED_COLUMNS and v not in ALL_PNL_COLUMNS:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        elif v in ALL_PNL_COLUMNS:
            cell.fill = PNL_HEADER_FILL
            cell.font = PNL_HEADER_FONT

    marker_ci = cm.get(marker_col) if marker_col else None
    if not marker_ci:
        marker_ci = cm.get("下注时间距开盘差(分，秒)") or cm.get("时间") or 1
    time_ci = marker_ci
    dir_ci = cm.get("净持仓方向")
    cyc_ci = cm.get("时间周期")

    sub_rows: set[int] = set()
    cycle_idx = 0
    prev_cyc = None
    for ri in range(2, ws.max_row + 1):
        tv = ws.cell(ri, time_ci).value if time_ci else None
        if tv == "【周期小计】":
            sub_rows.add(ri)
            continue
        if cyc_ci:
            cv = ws.cell(ri, cyc_ci).value
            if cv != prev_cyc:
                cycle_idx += 1
                prev_cyc = cv

    cycle_idx = 0
    prev_cyc = None
    for ri in range(2, ws.max_row + 1):
        is_sub = ri in sub_rows

        if not is_sub and cyc_ci:
            cv = ws.cell(ri, cyc_ci).value
            if cv != prev_cyc:
                cycle_idx += 1
                prev_cyc = cv

        if is_sub:
            for cell in ws[ri]:
                cell.fill = SUBTOTAL_FILL
                cell.font = SUBTOTAL_FONT
                cell.border = SUBTOTAL_BORDER

            for pc in ALL_PNL_COLUMNS:
                ci = cm.get(pc)
                if not ci:
                    continue
                cell = ws.cell(ri, ci)
                cell.number_format = "0.000"
                if isinstance(cell.value, (int, float)):
                    cell.font = PNL_POS_FONT if cell.value > 1e-10 else (
                        PNL_NEG_FONT if cell.value < -1e-10 else SUBTOTAL_FONT)

            if dir_ci:
                dc = ws.cell(ri, dir_ci)
                df_ = NET_DIRECTION_FILLS.get(dc.value)
                if df_:
                    dc.fill = df_
            continue

        cfill = CYCLE_FILLS[cycle_idx % 2]
        for cell in ws[ri]:
            cn = ws.cell(1, cell.column).value
            if cn in COL_FILLS:
                cell.fill = COL_FILLS[cn]
            elif cfill:
                cell.fill = cfill

        if dir_ci:
            dc = ws.cell(ri, dir_ci)
            df_ = NET_DIRECTION_FILLS.get(dc.value)
            if df_:
                dc.fill = df_

        for pc in ("Up已成交差价盈亏", "Down已成交差价盈亏", "浮动盈亏"):
            ci = cm.get(pc)
            if not ci:
                continue
            cell = ws.cell(ri, ci)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.000"
                cell.font = PNL_POS_FONT if cell.value > 1e-10 else (
                    PNL_NEG_FONT if cell.value < -1e-10 else Font())

    num_cols: set[int] = set()
    for cn, ci in cm.items():
        for ri in range(2, min(ws.max_row + 1, 20)):
            if ri in sub_rows:
                continue
            if isinstance(ws.cell(ri, ci).value, float):
                num_cols.add(ci)
                break
    for ci in num_cols:
        for ri in range(2, ws.max_row + 1):
            c = ws.cell(ri, ci)
            if isinstance(c.value, (int, float)):
                c.number_format = "0.000"
            c.alignment = Alignment(horizontal="right")

    percent_cols = (
        "同向成交价波动幅度(%)",
        "相对于加权均价的价格波动百分比",
        "持仓价值占比",
        "持仓价值加/减仓百分比",
    )
    for col_name in percent_cols:
        percent_ci = cm.get(col_name)
        if not percent_ci:
            continue
        for ri in range(2, ws.max_row + 1):
            c = ws.cell(ri, percent_ci)
            if isinstance(c.value, (int, float)):
                c.number_format = "0.0%"
            c.alignment = Alignment(horizontal="right")


# ═══════════════════════ CLI ═════════════════════════════════════════

def build_out(inp: Path, o: str | None) -> Path:
    if o:
        return Path(o)
    auto_name = f"{inp.stem}_with_accumulated_shares{inp.suffix}"
    if inp.parent.name == "raw" and inp.parent.parent.name == "data":
        out_dir = inp.parent.parent / "processed"
        out_dir.mkdir(parents=True, exist_ok=True)
        return out_dir / auto_name
    return inp.with_name(auto_name)


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Polymarket 持仓 + 盈亏分析")
    p.add_argument("--input", default="data/raw/polymarket_tracker_collection.xlsx")
    p.add_argument("--output", default=None)
    p.add_argument("--sheet", default=None)
    p.add_argument("--price-col", default=None)
    p.add_argument("--qty-col", default=None)
    p.add_argument("--outcome-col", default=None)
    p.add_argument("--action-col", default=None)
    p.add_argument("--coin-col", default=None)
    p.add_argument("--cycle-col", default=None)
    p.add_argument("--group-cols", type=split_group_cols, default=None)
    p.add_argument("--config", default="configs/strategy.yaml")
    p.add_argument("--position-value-denominator", type=float, default=None)
    p.add_argument(
        "--winner",
        choices=["up", "down"],
        default=None,
        help="覆盖周期 winner 判定（来自引擎结算结果），不传则按用户最后成交价推断。",
    )
    return p.parse_args()


def _resolve_position_value_denominator(args: argparse.Namespace) -> float:
    explicit = getattr(args, "position_value_denominator", None)
    if explicit is not None:
        value = float(explicit)
        if value <= 1e-12:
            raise ValueError("--position-value-denominator 必须 > 0")
        return value

    cfg_path = Path(getattr(args, "config", "configs/strategy.yaml"))
    if cfg_path.exists():
        data = yaml.safe_load(cfg_path.read_text(encoding="utf-8")) or {}
        if isinstance(data, dict):
            position = data.get("position", {})
            if isinstance(position, dict) and position.get("max_position_value") is not None:
                value = float(position["max_position_value"])
                if value > 1e-12:
                    return value

    raise ValueError("未能解析持仓占比分母，请传 --position-value-denominator 或提供有效 --config")


def main() -> int:
    args = parse_args()
    args.position_value_denominator = _resolve_position_value_denominator(args)
    inp = Path(args.input)
    out = build_out(inp, args.output)
    if not inp.exists():
        print(f"未找到: {inp}", file=sys.stderr)
        return 1
    try:
        xls = pd.ExcelFile(inp)
        tgts = [args.sheet] if args.sheet else list(xls.sheet_names)
        if args.sheet and args.sheet not in xls.sheet_names:
            raise ValueError(f"工作表 `{args.sheet}` 不存在。")
        frames: dict[str, pd.DataFrame] = {}
        sums: list[tuple[str, dict]] = []
        for sn in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sn)
            if sn in tgts:
                proc, sm = compute(df, args, winner_override=args.winner)
                frames[sn] = proc
                sums.append((sn, sm))
            else:
                frames[sn] = df
        if not sums:
            raise RuntimeError("未生成结果。")
        with pd.ExcelWriter(out, engine="openpyxl") as w:
            for sn in xls.sheet_names:
                frames[sn].to_excel(w, sheet_name=sn, index=False)
                sm_row = next((s for name, s in sums if name == sn), None)
                format_ws(w.sheets[sn], sm_row.get("marker_col") if sm_row else None)
        print(f"完成: {out}")
        for sn, sm in sums:
            print(f"  [{sn}] 周期={sm['cycle_col']} 份数={sm['qty_col']} 方向={sm['outcome_col']} 操作={sm['action_col']}")
        return 0
    except Exception as exc:
        print(f"失败: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
