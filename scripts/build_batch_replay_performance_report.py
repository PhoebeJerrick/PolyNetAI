from __future__ import annotations

import argparse
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
import yaml

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "src"
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

# 分周期执行交易流水（原 BTC 工作表）
TRACKER_STYLE_SHEET = "分周期执行交易流水"

# 与 write_batch_trade_process_zh 决策表列顺序一致
TRADE_PROCESS_PREF_COLS = (
    "timestamp",
    "market_price",
    "market_outcome",
    "selected_rule",
    "selected_outcome",
    "selected_action",
    "selected_shares",
    "risk_status",
    "risk_reason",
    "executed",
    "submitted",
    "confirmed",
    "fill_price",
    "fill_fee",
    "cycle_net_profit",
    "account_cash",
)

from polynet_ai.reporting.excel_export import (
    SAME_OUTCOME_PRICE_MOVE_COL,
    absolute_cycle_open_timestamp_utc_for_cycle_id,
    _format_elapsed,
    compute_same_outcome_price_move_pct,
    get_version_tag,
    infer_decision_reason,
)

_TRADE_PROCESS_ELAPSED_COL = "下注时间距开盘差(分,秒)"
DEFAULT_POSITION_VALUE_DENOMINATOR = 85.0
DEFAULT_PHASE_END_SECONDS: tuple[float, float, float] = (70.0, 160.0, 240.0)


def _normalize_cycle_slug(value: object) -> str:
    """cycle_slug 仅保留周期目录名，避免在报表中出现绝对路径。"""
    text = str(value or "").strip()
    if not text:
        return ""
    return text.replace("\\", "/").rstrip("/").split("/")[-1]


def _cycle_starts_from_full_decision_frame(full: pd.DataFrame) -> pd.DataFrame:
    """``full`` 须含 ``market_id``、``_pid``、``timestamp``（已 ``to_datetime(..., utc=True)``）。

    返回各 (market_id, _pid) 的绝对开盘（slug epoch）UTC 时间；无法解析 slug 时回退为该组最早一条数据时间。
    """
    starts_min = full.groupby(["market_id", "_pid"], dropna=False)["timestamp"].min().reset_index(
        name="_cycle_start_fallback"
    )
    key_open = full[["market_id", "_pid"]].drop_duplicates().copy()
    key_open["_from_slug"] = key_open["_pid"].map(absolute_cycle_open_timestamp_utc_for_cycle_id)
    key_open["_from_slug"] = pd.to_datetime(key_open["_from_slug"], errors="coerce", utc=True)
    starts_min["_cycle_start_fallback"] = pd.to_datetime(
        starts_min["_cycle_start_fallback"], errors="coerce", utc=True
    )
    key_open = key_open.merge(starts_min, on=["market_id", "_pid"], how="left")
    # 优先使用更早的时间戳，避免 slug 解析结果偏晚时把有效成交排除在周期窗口外。
    # pandas 2/3：两列 min(axis=1) 可能混用 tz-aware / NaT，与 decision 时间戳相减前统一为 UTC。
    key_open["_cycle_start"] = pd.to_datetime(
        key_open[["_from_slug", "_cycle_start_fallback"]].min(axis=1),
        utc=True,
    )
    return key_open[["market_id", "_pid", "_cycle_start"]]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="根据批量回放结果生成中文总绩效报告")
    parser.add_argument(
        "--input-dir",
        required=True,
        help="可传抓取目录或 batch_replay_outputs 目录，例如 artifacts/live/ws_capture_btc_10cycles",
    )
    parser.add_argument(
        "--output-dir",
        default=None,
        help="默认输出到解析后的 batch_replay_outputs 目录",
    )
    parser.add_argument("--config", default="configs/strategy.yaml", help="策略配置文件，用于读取 position.max_position_value")
    parser.add_argument("--position-value-denominator", type=float, default=None, help="持仓价值占比分母，优先级高于 --config")
    return parser.parse_args()


def _resolve_batch_replay_dir(input_dir: str | Path) -> Path:
    path = Path(input_dir)
    if not path.exists():
        raise FileNotFoundError(f"未找到输入目录: {path}")
    nested = path / "batch_replay_outputs"
    report_markers = ("batch_replay_summary.csv",)
    report_globs = ("*batch_replay_performance_report*.xlsx",)
    has_report_in_path = any((path / name).exists() for name in report_markers) or any(
        next(path.glob(pattern), None) is not None for pattern in report_globs
    )
    has_report_in_nested = any((nested / name).exists() for name in report_markers) or any(
        next(nested.glob(pattern), None) is not None for pattern in report_globs
    )
    if has_report_in_path:
        return path
    if has_report_in_nested:
        return nested
    if nested.exists():
        return nested
    return path


def _discover_cycle_result_dirs(batch_dir: Path) -> list[Path]:
    return sorted(
        path
        for path in batch_dir.iterdir()
        if path.is_dir()
        and (path / "cycles.csv").exists()
        and (path / "decisions.csv").exists()
        and (path / "metrics.csv").exists()
    )


def resolve_position_value_denominator(
    *,
    config_path: str | Path = "configs/strategy.yaml",
    explicit: float | None = None,
) -> float:
    """统一持仓价值占比分母：优先显式参数，否则读取 position.max_position_value。"""
    if explicit is not None:
        value = float(explicit)
        if value <= 1e-12:
            raise ValueError("position_value_denominator 必须 > 0")
        return value

    cfg_path = Path(config_path)
    if cfg_path.exists():
        data = yaml.safe_load(cfg_path.read_text(encoding="utf-8")) or {}
        if isinstance(data, dict):
            position = data.get("position", {})
            if isinstance(position, dict) and position.get("max_position_value") is not None:
                value = float(position["max_position_value"])
                if value > 1e-12:
                    return value
    return DEFAULT_POSITION_VALUE_DENOMINATOR


def resolve_phase_end_seconds(
    *,
    config_path: str | Path = "configs/strategy.yaml",
) -> tuple[float, float, float]:
    """从配置读取阶段边界（严格递增），失败时回退默认值。"""
    cfg_path = Path(config_path)
    if not cfg_path.exists():
        return DEFAULT_PHASE_END_SECONDS
    try:
        data = yaml.safe_load(cfg_path.read_text(encoding="utf-8")) or {}
    except Exception:
        return DEFAULT_PHASE_END_SECONDS
    if not isinstance(data, dict):
        return DEFAULT_PHASE_END_SECONDS
    cycle = data.get("cycle", {})
    if not isinstance(cycle, dict):
        return DEFAULT_PHASE_END_SECONDS
    try:
        e1 = float(cycle.get("phase_end_seconds_1", DEFAULT_PHASE_END_SECONDS[0]))
        e2 = float(cycle.get("phase_end_seconds_2", DEFAULT_PHASE_END_SECONDS[1]))
        e3 = float(cycle.get("phase_end_seconds_3", DEFAULT_PHASE_END_SECONDS[2]))
    except (TypeError, ValueError):
        return DEFAULT_PHASE_END_SECONDS
    if not (0.0 < e1 < e2 < e3):
        return DEFAULT_PHASE_END_SECONDS
    return (e1, e2, e3)


def _compute_max_drawdown(profits: list[float]) -> float:
    if not profits:
        return 0.0
    cumulative = 0.0
    peak = 0.0
    max_drawdown = 0.0
    for profit in profits:
        cumulative += float(profit)
        peak = max(peak, cumulative)
        max_drawdown = max(max_drawdown, peak - cumulative)
    return max_drawdown


def _load_summary(batch_dir: Path, cycle_dirs: list[Path]) -> pd.DataFrame:
    summary_csv = batch_dir / "batch_replay_summary.csv"
    if summary_csv.exists():
        summary_df = pd.read_csv(summary_csv)
        if not summary_df.empty:
            if "cycle_slug" in summary_df.columns:
                summary_df["cycle_slug"] = summary_df["cycle_slug"].map(_normalize_cycle_slug)
            return summary_df.sort_values("cycle_slug").reset_index(drop=True)

    rows: list[dict[str, object]] = []
    for cycle_dir in cycle_dirs:
        cycle_slug = cycle_dir.name
        metrics = pd.read_csv(cycle_dir / "metrics.csv").iloc[0].to_dict()
        cycles = pd.read_csv(cycle_dir / "cycles.csv").iloc[0].to_dict()
        rows.append(
            {
                "cycle_slug": cycle_slug,
                "executed_trades": metrics.get("executed_trades", 0),
                "accepted_signals": metrics.get("accepted_signals", 0),
                "blocked_signals": metrics.get("blocked_signals", 0),
                "total_net_profit": metrics.get("total_net_profit", 0.0),
                "total_fees": metrics.get("total_fees", 0.0),
                "winner": cycles.get("winner", ""),
                "account_cash": cycles.get("account_cash", None),
            }
        )
    summary_df = pd.DataFrame(rows)
    if not summary_df.empty:
        summary_df = summary_df.sort_values("cycle_slug").reset_index(drop=True)
    return summary_df


def _load_cycle_frames(cycle_dirs: list[Path]) -> tuple[pd.DataFrame, pd.DataFrame]:
    cycle_rows: list[pd.DataFrame] = []
    decision_rows: list[pd.DataFrame] = []
    for cycle_dir in cycle_dirs:
        cycle_slug = cycle_dir.name
        cycles_df = pd.read_csv(cycle_dir / "cycles.csv").copy()
        decisions_df = pd.read_csv(cycle_dir / "decisions.csv").copy()
        cycles_df["cycle_slug"] = cycle_slug
        decisions_df["cycle_slug"] = cycle_slug
        cycle_rows.append(cycles_df)
        decision_rows.append(decisions_df)
    cycle_df = pd.concat(cycle_rows, ignore_index=True) if cycle_rows else pd.DataFrame()
    decision_df = pd.concat(decision_rows, ignore_index=True) if decision_rows else pd.DataFrame()
    if not cycle_df.empty and "cycle_slug" in cycle_df.columns:
        cycle_df["cycle_slug"] = cycle_df["cycle_slug"].map(_normalize_cycle_slug)
    if not cycle_df.empty and "cycle_id" in cycle_df.columns:
        cycle_df["cycle_id"] = cycle_df["cycle_id"].map(_normalize_cycle_slug)
    if not decision_df.empty:
        if "cycle_slug" in decision_df.columns:
            decision_df["cycle_slug"] = decision_df["cycle_slug"].map(_normalize_cycle_slug)
        if "cycle_id" in decision_df.columns:
            decision_df["cycle_id"] = decision_df["cycle_id"].map(_normalize_cycle_slug)
    return cycle_df, decision_df


def _summarize_direction_distribution(decision_df: pd.DataFrame) -> pd.DataFrame:
    if decision_df.empty or "executed" not in decision_df.columns:
        return pd.DataFrame(columns=["selected_outcome", "selected_action", "trades", "shares"])
    executed_df = decision_df[decision_df["executed"].fillna(False).astype(bool)].copy()
    if executed_df.empty:
        return pd.DataFrame(columns=["selected_outcome", "selected_action", "trades", "shares"])
    executed_df["selected_outcome"] = executed_df["selected_outcome"].fillna("").astype(str)
    executed_df["selected_action"] = executed_df["selected_action"].fillna("").astype(str)
    executed_df["selected_shares"] = pd.to_numeric(executed_df["selected_shares"], errors="coerce").fillna(0.0)
    summary = (
        executed_df.groupby(["selected_outcome", "selected_action"], dropna=False)
        .agg(trades=("executed", "size"), shares=("selected_shares", "sum"))
        .reset_index()
        .sort_values(["selected_outcome", "selected_action"])
        .reset_index(drop=True)
    )
    return summary


def _value_counts_frame(series: pd.Series, name: str) -> pd.DataFrame:
    if series.empty:
        return pd.DataFrame(columns=[name, "count"])
    counts = (
        series.fillna("unknown")
        .astype(str)
        .value_counts(dropna=False)
        .rename_axis(name)
        .reset_index(name="count")
    )
    return counts


def _infer_cycle_length_seconds(cycle_slug: object) -> int:
    text = str(cycle_slug or "").lower()
    m = re.search(r"-(\d+)m-", text)
    if m:
        return int(m.group(1)) * 60
    return 300


def _phase_label_for_elapsed_seconds(
    elapsed_seconds: float,
    cycle_length_seconds: int,
    *,
    phase_end_seconds: tuple[float, float, float] = DEFAULT_PHASE_END_SECONDS,
) -> str:
    """按周期进度切分四段，生成轻量阶段标签（P1~P4，便于大数据筛选）。"""
    if cycle_length_seconds <= 0:
        cycle_length_seconds = 300
    if pd.isna(elapsed_seconds):
        return ""
    sec = max(0.0, min(float(elapsed_seconds), float(cycle_length_seconds)))
    e1, e2, e3 = phase_end_seconds
    if sec <= e1:
        return "P1"
    if sec <= e2:
        return "P2"
    if sec <= e3:
        return "P3"
    return "P4"


def _is_up_outcome(value: object) -> bool:
    text = str(value or "").strip().lower()
    return text in {"up", "yes"}


def _is_down_outcome(value: object) -> bool:
    text = str(value or "").strip().lower()
    return text in {"down", "no"}


def _build_cycle_price_frame(
    decision_df: pd.DataFrame,
    cycle_df: pd.DataFrame,
    *,
    complete_only: bool,
) -> pd.DataFrame:
    """构建每条行情事件对应的 Up/Down 价格明细，并补齐周期内秒数。"""
    columns = [
        "cycle_slug",
        "timestamp",
        "周期内秒数",
        "下注时间距开盘差(分,秒)",
        "market_outcome",
        "market_price",
        "Up价格",
        "Down价格",
        "selected_outcome",
        "selected_action",
        "executed",
    ]
    if decision_df.empty:
        return pd.DataFrame(columns=columns)

    dec = decision_df.copy()
    if "cycle_slug" not in dec.columns:
        dec["cycle_slug"] = dec.get("cycle_id", pd.Series(dtype=object)).fillna("").astype(str)
    else:
        dec["cycle_slug"] = dec["cycle_slug"].fillna(dec.get("cycle_id", "")).astype(str)

    if complete_only and not cycle_df.empty:
        complete_cycle_slugs = set(
            cycle_df.get("cycle_slug", cycle_df.get("cycle_id", pd.Series(dtype=object)))
            .fillna("")
            .astype(str)
            .tolist()
        )
        dec = dec[dec["cycle_slug"].isin(complete_cycle_slugs)]
    if dec.empty:
        return pd.DataFrame(columns=columns)

    if "timestamp" not in dec.columns:
        dec["timestamp"] = pd.NaT
    dec["timestamp"] = pd.to_datetime(dec["timestamp"], errors="coerce", utc=True)

    if "market_id" not in dec.columns or dec["market_id"].isna().all():
        if not cycle_df.empty and "cycle_slug" in cycle_df.columns and "market_id" in cycle_df.columns:
            dec = dec.merge(
                cycle_df.drop_duplicates(subset=["cycle_slug"], keep="first")[["cycle_slug", "market_id"]],
                on="cycle_slug",
                how="left",
            )
    dec["market_id"] = dec.get("market_id", pd.Series(dtype=object)).fillna("").astype(str)
    if "cycle_id" in dec.columns:
        dec["_pid"] = dec["cycle_id"].fillna(dec.get("cycle_slug", "")).astype(str)
    else:
        dec["_pid"] = dec.get("cycle_slug", pd.Series(dtype=object)).fillna("").astype(str)

    starts = _cycle_starts_from_full_decision_frame(dec)
    dec = dec.merge(starts, on=["market_id", "_pid"], how="left")
    dec["_cycle_start"] = pd.to_datetime(dec["_cycle_start"], errors="coerce", utc=True)

    dec["周期内秒数"] = (dec["timestamp"] - dec["_cycle_start"]).dt.total_seconds()
    dec["周期内秒数"] = pd.to_numeric(dec["周期内秒数"], errors="coerce")
    dec.loc[dec["周期内秒数"] < 0, "周期内秒数"] = 0.0
    dec["下注时间距开盘差(分,秒)"] = [
        _format_elapsed(ts, t0) for ts, t0 in zip(dec["timestamp"], dec["_cycle_start"])
    ]

    market_price = pd.to_numeric(dec.get("market_price", pd.Series(dtype=object)), errors="coerce")
    if market_price.isna().all():
        market_price = pd.to_numeric(dec.get("fill_price", pd.Series(dtype=object)), errors="coerce")
    outcome = dec.get("market_outcome", pd.Series(dtype=object)).fillna("").astype(str)

    up_price = pd.Series(pd.NA, index=dec.index, dtype="Float64")
    down_price = pd.Series(pd.NA, index=dec.index, dtype="Float64")

    up_mask = outcome.map(_is_up_outcome)
    down_mask = outcome.map(_is_down_outcome)
    up_price.loc[up_mask] = market_price.loc[up_mask]
    down_price.loc[up_mask] = 1.0 - market_price.loc[up_mask]
    down_price.loc[down_mask] = market_price.loc[down_mask]
    up_price.loc[down_mask] = 1.0 - market_price.loc[down_mask]

    result = pd.DataFrame(
        {
            "cycle_slug": dec["cycle_slug"],
            "timestamp": dec["timestamp"].dt.tz_convert("UTC").dt.tz_localize(None),
            "周期内秒数": dec["周期内秒数"],
            "下注时间距开盘差(分,秒)": dec["下注时间距开盘差(分,秒)"],
            "market_outcome": outcome,
            "market_price": market_price,
            "Up价格": up_price,
            "Down价格": down_price,
            "selected_outcome": dec.get("selected_outcome", pd.Series(dtype=object)).fillna("").astype(str),
            "selected_action": dec.get("selected_action", pd.Series(dtype=object)).fillna("").astype(str),
            "executed": dec.get("executed", pd.Series(dtype=object)),
        }
    )
    result = result[result["timestamp"].notna()].copy()
    result = result.sort_values(["cycle_slug", "timestamp"], kind="mergesort").reset_index(drop=True)
    return result[columns]


def _short_cycle_label_for_chart(cycle_slug: object) -> str:
    """图表标题用短名：路径取最后一段；若该段以 ``-`` 分隔且末段为纯数字（如开盘 epoch）则只保留该数字。"""
    s = str(cycle_slug or "").strip()
    if not s:
        return ""
    base = s.replace("\\", "/").rstrip("/").split("/")[-1]
    if not base:
        return s
    tail = base.split("-")[-1]
    if tail.isdigit():
        return tail
    return base


def _append_cycle_price_charts(
    ws,
    cycle_price_df: pd.DataFrame,
    *,
    title_prefix: str,
) -> None:
    """在同一 sheet 右侧追加每个周期的 Up/Down 价格折线图。"""
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.series import SeriesLabel

    if cycle_price_df.empty:
        return

    col_map = {name: idx + 1 for idx, name in enumerate(cycle_price_df.columns)}
    cycle_col = col_map["cycle_slug"]
    sec_col = col_map["周期内秒数"]
    up_col = col_map["Up价格"]
    down_col = col_map["Down价格"]

    row_cursor = 2
    chart_idx = 0
    max_row = len(cycle_price_df) + 1
    while row_cursor <= max_row:
        cycle_value = ws.cell(row=row_cursor, column=cycle_col).value
        row_end = row_cursor
        while row_end + 1 <= max_row and ws.cell(row=row_end + 1, column=cycle_col).value == cycle_value:
            row_end += 1

        chart = LineChart()
        chart.title = f"{title_prefix}-{_short_cycle_label_for_chart(cycle_value)}"
        chart.y_axis.title = "价格"
        chart.x_axis.title = "周期内秒数"
        chart.height = 6.5
        chart.width = 11.5

        up_ref = Reference(ws, min_col=up_col, min_row=row_cursor, max_row=row_end)
        down_ref = Reference(ws, min_col=down_col, min_row=row_cursor, max_row=row_end)
        x_ref = Reference(ws, min_col=sec_col, min_row=row_cursor, max_row=row_end)

        chart.add_data(up_ref, titles_from_data=False)
        chart.add_data(down_ref, titles_from_data=False)
        chart.set_categories(x_ref)
        if len(chart.series) >= 2:
            chart.series[0].title = SeriesLabel(v=str(ws.cell(row=1, column=up_col).value or "Up价格"))
            chart.series[1].title = SeriesLabel(v=str(ws.cell(row=1, column=down_col).value or "Down价格"))

        anchor_row = 2 + chart_idx * 15
        ws.add_chart(chart, f"N{anchor_row}")

        chart_idx += 1
        row_cursor = row_end + 1


def _write_cycle_price_sheet(
    writer: pd.ExcelWriter,
    *,
    sheet_name: str,
    cycle_price_df: pd.DataFrame,
    chart_title_prefix: str,
    include_charts: bool = True,
) -> None:
    if cycle_price_df.empty:
        pd.DataFrame({"说明": ["无可用价格数据，未生成分周期价格图表。"]}).to_excel(
            writer, sheet_name=sheet_name, index=False
        )
        return
    cycle_price_df.to_excel(writer, sheet_name=sheet_name, index=False)
    if include_charts:
        ws = writer.sheets[sheet_name]
        _append_cycle_price_charts(ws, cycle_price_df, title_prefix=chart_title_prefix)


def _resolve_trade_process_chart_mode(
    *,
    cycle_df: pd.DataFrame,
    decision_df: pd.DataFrame,
) -> bool:
    """决定交易过程 Excel 是否生成图表。大批量默认关闭以缩短耗时。"""
    import os

    raw = str(os.getenv("POLYNET_TRADE_PROCESS_CHARTS", "auto")).strip().lower()
    if raw in {"1", "true", "yes", "on"}:
        return True
    if raw in {"0", "false", "no", "off"}:
        return False

    cycle_count = int(cycle_df.get("cycle_slug", pd.Series(dtype=object)).nunique()) if not cycle_df.empty else 0
    decision_rows = int(len(decision_df))
    # auto: 大规模任务默认禁用图表（图表构建是 openpyxl 最耗时部分）
    return not (cycle_count >= 25 or decision_rows >= 120_000)


def _apply_basic_sheet_style(
    writer: pd.ExcelWriter,
    *,
    skip_sheet_titles: frozenset[str] | None = None,
) -> None:
    """在同一次写盘中完成首行冻结/加粗，避免再次 load_workbook 全量重写。"""
    from openpyxl.styles import Font

    skip = skip_sheet_titles or frozenset()
    for name, ws in writer.sheets.items():
        if name in skip or ws.max_row == 0:
            continue
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)


def summarize_tail_window_executions(
    decision_df: pd.DataFrame,
    cycle_df: pd.DataFrame,
    *,
    last_minute_seconds: int = 30,
) -> tuple[pd.DataFrame, list[tuple[str, object]]]:
    """
    按周期汇总「周期末 last_minute_seconds 内」的已执行成交：笔数、现金流净额（不含结算）、手续费。
    周期长度从 cycle_slug 中的 ``-Xm-`` 推断（如 5m -> 300s），否则 300s。
    """
    empty_overview: list[tuple[str, object]] = [
        (
            "尾盘窗口说明",
            f"周期末最后 {last_minute_seconds}s（周期长度从 slug 中 -Xm- 推断，否则 300s）",
        ),
        ("尾盘已执行成交笔数(合计)", 0),
        ("尾盘成交现金流净额(不含结算)", 0.0),
        ("尾盘手续费合计", 0.0),
    ]
    if decision_df.empty or "cycle_slug" not in decision_df.columns:
        return pd.DataFrame(), empty_overview

    full = decision_df.copy()
    full["timestamp"] = pd.to_datetime(full.get("timestamp"), errors="coerce", utc=True)
    if "market_id" not in full.columns or full["market_id"].isna().all():
        if not cycle_df.empty and "cycle_slug" in full.columns and "market_id" in cycle_df.columns:
            full = full.merge(
                cycle_df.drop_duplicates(subset=["cycle_slug"], keep="first")[["cycle_slug", "market_id"]],
                on="cycle_slug",
                how="left",
            )
    full["market_id"] = full.get("market_id", pd.Series(dtype=object)).fillna("").astype(str)
    if "cycle_id" in full.columns:
        full["_pid"] = full["cycle_id"].fillna(full.get("cycle_slug", "")).astype(str)
    else:
        full["_pid"] = full.get("cycle_slug", pd.Series(dtype=object)).fillna("").astype(str)
    starts = _cycle_starts_from_full_decision_frame(full)

    dec = decision_df.copy()
    if "executed" in dec.columns:
        dec = dec[dec["executed"].fillna(False).astype(bool)]
    if dec.empty:
        return pd.DataFrame(), empty_overview

    if "market_id" not in dec.columns or dec["market_id"].isna().all():
        if not cycle_df.empty and "cycle_slug" in dec.columns and "market_id" in cycle_df.columns:
            dec = dec.merge(
                cycle_df.drop_duplicates(subset=["cycle_slug"], keep="first")[["cycle_slug", "market_id"]],
                on="cycle_slug",
                how="left",
            )
    dec["market_id"] = dec.get("market_id", pd.Series(dtype=object)).fillna("").astype(str)
    if "cycle_id" in dec.columns:
        dec["_pid"] = dec["cycle_id"].fillna(dec.get("cycle_slug", "")).astype(str)
    else:
        dec["_pid"] = dec.get("cycle_slug", pd.Series(dtype=object)).fillna("").astype(str)
    dec["timestamp"] = pd.to_datetime(dec.get("timestamp"), errors="coerce", utc=True)
    dec = dec.merge(starts, on=["market_id", "_pid"], how="left")
    dec = dec[dec["_cycle_start"].notna() & dec["timestamp"].notna()]
    if dec.empty:
        return pd.DataFrame(), empty_overview

    dec["_cycle_len"] = dec["cycle_slug"].map(_infer_cycle_length_seconds)
    dec["_elapsed_sec"] = (dec["timestamp"] - dec["_cycle_start"]).dt.total_seconds()
    tail_cut = dec["_cycle_len"] - int(last_minute_seconds)
    tail = dec[dec["_elapsed_sec"] >= tail_cut].copy()
    if tail.empty:
        return pd.DataFrame(), empty_overview

    sh = pd.to_numeric(tail["selected_shares"], errors="coerce").fillna(0.0)
    px = pd.to_numeric(tail["fill_price"], errors="coerce").fillna(0.0)
    fee = pd.to_numeric(tail.get("fill_fee", 0.0), errors="coerce").fillna(0.0)
    act = tail["selected_action"].fillna("").astype(str).str.lower()
    cash = pd.Series(0.0, index=tail.index, dtype=float)
    m_buy = act == "buy"
    m_sell = act == "sell"
    cash.loc[m_buy] = -(sh[m_buy] * px[m_buy] + fee[m_buy])
    cash.loc[m_sell] = sh[m_sell] * px[m_sell] - fee[m_sell]
    tail["_cash_flow"] = cash

    per_cycle = (
        tail.groupby("cycle_slug", dropna=False)
        .agg(
            tail_executed_trades=("selected_shares", "size"),
            tail_cash_flow_net=("_cash_flow", "sum"),
            tail_fees=("fill_fee", lambda s: pd.to_numeric(s, errors="coerce").fillna(0.0).sum()),
        )
        .reset_index()
        .sort_values("cycle_slug", kind="mergesort")
        .reset_index(drop=True)
    )

    overview_rows: list[tuple[str, object]] = [
        (
            "尾盘窗口说明",
            f"周期末最后 {last_minute_seconds}s（周期长度从 slug 中 -Xm- 推断，否则 300s）",
        ),
        ("尾盘已执行成交笔数(合计)", int(len(tail))),
        ("尾盘成交现金流净额(不含结算)", float(tail["_cash_flow"].sum())),
        ("尾盘手续费合计", float(pd.to_numeric(tail["fill_fee"], errors="coerce").fillna(0.0).sum())),
    ]
    return per_cycle, overview_rows


def _tracker_compute_args(
    *,
    position_value_denominator: float = DEFAULT_POSITION_VALUE_DENOMINATOR,
) -> argparse.Namespace:
    """与 analyze_polymarket_tracker.compute 默认列名一致。"""
    return argparse.Namespace(
        group_cols=None,
        coin_col="市场标题",
        cycle_col="时间周期",
        qty_col="投注份数",
        outcome_col="结果代币类型",
        action_col="操作方向",
        price_col="成交价格",
        position_value_denominator=position_value_denominator,
    )


def _batch_replay_decisions_to_tracker_input(
    cycle_df: pd.DataFrame,
    decision_df: pd.DataFrame,
    *,
    phase_end_seconds: tuple[float, float, float] = DEFAULT_PHASE_END_SECONDS,
) -> pd.DataFrame:
    """
    将批量回放的 cycle / decision 合并为 analyze_polymarket_tracker.compute 所需的宽表。
    与 polymarket_tracker_collection_with_accumulated_shares_v5.xlsx 使用同一套 compute/format_ws 逻辑。
    """
    if decision_df.empty:
        return pd.DataFrame()
    dec = decision_df.copy()
    if "executed" in dec.columns:
        dec = dec[dec["executed"].fillna(False).astype(bool)]
    if dec.empty:
        return pd.DataFrame()

    dec["_sort"] = range(len(dec))

    full = decision_df.copy()
    full["timestamp"] = pd.to_datetime(full.get("timestamp"), errors="coerce", utc=True)
    if "market_id" not in full.columns or full["market_id"].isna().all():
        if not cycle_df.empty and "cycle_slug" in full.columns and "market_id" in cycle_df.columns:
            full = full.merge(
                cycle_df.drop_duplicates(subset=["cycle_slug"], keep="first")[["cycle_slug", "market_id"]],
                on="cycle_slug",
                how="left",
            )
    full["market_id"] = full.get("market_id", pd.Series(dtype=object)).fillna("").astype(str)
    if "cycle_id" in full.columns:
        full["_pid"] = full["cycle_id"].fillna(full.get("cycle_slug", "")).astype(str)
    else:
        full["_pid"] = full.get("cycle_slug", pd.Series(dtype=object)).fillna("").astype(str)

    starts = _cycle_starts_from_full_decision_frame(full)

    if "market_id" not in dec.columns or dec["market_id"].isna().all():
        if not cycle_df.empty and "cycle_slug" in dec.columns and "market_id" in cycle_df.columns:
            dec = dec.merge(
                cycle_df.drop_duplicates(subset=["cycle_slug"], keep="first")[["cycle_slug", "market_id"]],
                on="cycle_slug",
                how="left",
            )
    dec["market_id"] = dec.get("market_id", pd.Series(dtype=object)).fillna("").astype(str)
    if "cycle_id" in dec.columns:
        dec["_pid"] = dec["cycle_id"].fillna(dec.get("cycle_slug", "")).astype(str)
    else:
        dec["_pid"] = dec.get("cycle_slug", pd.Series(dtype=object)).fillna("").astype(str)

    dec["timestamp"] = pd.to_datetime(dec.get("timestamp"), errors="coerce", utc=True)
    dec = dec.merge(starts, on=["market_id", "_pid"], how="left")

    if "cycle_id" in dec.columns:
        period = dec["cycle_id"].fillna(dec.get("cycle_slug", "")).astype(str)
    else:
        period = dec["_pid"].astype(str)

    shares = pd.to_numeric(dec.get("selected_shares", pd.Series(0.0, index=dec.index)), errors="coerce").fillna(0.0)
    fill_px = pd.to_numeric(dec.get("fill_price", pd.Series(0.0, index=dec.index)), errors="coerce").fillna(0.0)
    same_move = compute_same_outcome_price_move_pct(
        dec,
        outcome_col="selected_outcome",
        price_col="fill_price",
        group_cols=["market_id", "_pid"],
    )
    decision_reason = infer_decision_reason(dec)

    elapsed_seconds = (dec["timestamp"] - dec["_cycle_start"]).dt.total_seconds()
    cycle_len = dec["_pid"].map(_infer_cycle_length_seconds).astype(float)
    phase_col = [
        _phase_label_for_elapsed_seconds(
            el,
            int(cl) if pd.notna(cl) else 300,
            phase_end_seconds=phase_end_seconds,
        )
        for el, cl in zip(elapsed_seconds, cycle_len)
    ]

    raw = pd.DataFrame(
        {
            "下注时间距开盘差(分，秒)": [
                _format_elapsed(ts, t0) for ts, t0 in zip(dec["timestamp"], dec["_cycle_start"])
            ],
            "市场标题": dec["market_id"],
            "时间周期": period,
            "阶段": phase_col,
            "结果代币类型": dec.get("selected_outcome", pd.Series(dtype=object)).fillna("").astype(str),
            "操作方向": dec.get("selected_action", pd.Series(dtype=object)).fillna("").astype(str),
            "决策原因": decision_reason,
            "投注份数": shares,
            "USDT价值": shares * fill_px,
            "成交价格": fill_px,
            SAME_OUTCOME_PRICE_MOVE_COL: same_move,
        }
    )
    raw["_sort"] = dec["_sort"]
    raw = raw.sort_values("_sort").drop(columns=["_sort"])
    return raw.reset_index(drop=True)


def _winner_label(value: object) -> str:
    text = str(value or "").strip().lower()
    if text == "up":
        return "Up"
    if text == "down":
        return "Down"
    return ""


def _align_tracker_subtotals_with_cycle_snapshot(
    tracker_df: pd.DataFrame,
    cycle_df: pd.DataFrame,
) -> pd.DataFrame:
    """
    用 cycle_df 的周期结算结果回填 BTC 表的【周期小计】行，避免与周期快照口径不一致。
    """
    if tracker_df.empty or cycle_df.empty:
        return tracker_df
    marker_col = "下注时间距开盘差(分，秒)"
    cycle_col = "时间周期"
    if marker_col not in tracker_df.columns or cycle_col not in tracker_df.columns:
        return tracker_df
    if "cycle_id" not in cycle_df.columns:
        return tracker_df

    out = tracker_df.copy()
    subtotal_mask = out[marker_col].astype(str) == "【周期小计】"
    if not subtotal_mask.any():
        return out

    cycle_lookup = cycle_df.drop_duplicates(subset=["cycle_id"], keep="last").set_index("cycle_id")

    def _get(cycle_id: str, col: str) -> object:
        if col not in cycle_lookup.columns:
            return None
        try:
            return cycle_lookup.at[cycle_id, col]
        except KeyError:
            return None

    for idx in out.index[subtotal_mask]:
        cycle_id = str(out.at[idx, cycle_col])
        if not cycle_id:
            continue
        if "最终Winner方向" in out.columns:
            out.at[idx, "最终Winner方向"] = _winner_label(_get(cycle_id, "winner"))
        if "未平仓UP盈亏" in out.columns:
            value = _get(cycle_id, "unrealized_up_pnl")
            if value is not None:
                out.at[idx, "未平仓UP盈亏"] = value
        if "未平仓Down盈亏" in out.columns:
            value = _get(cycle_id, "unrealized_down_pnl")
            if value is not None:
                out.at[idx, "未平仓Down盈亏"] = value
        if "周期净利润" in out.columns:
            value = _get(cycle_id, "cycle_net_profit")
            if value is not None:
                out.at[idx, "周期净利润"] = value
        if "Up已成交差价盈亏" in out.columns:
            value = _get(cycle_id, "up_realized_pnl")
            if value is not None:
                out.at[idx, "Up已成交差价盈亏"] = value
        if "Down已成交差价盈亏" in out.columns:
            value = _get(cycle_id, "down_realized_pnl")
            if value is not None:
                out.at[idx, "Down已成交差价盈亏"] = value
    return out


def _append_tracker_style_sheet(
    writer: pd.ExcelWriter,
    raw_input: pd.DataFrame,
    cycle_df: pd.DataFrame,
    *,
    position_value_denominator: float = DEFAULT_POSITION_VALUE_DENOMINATOR,
) -> None:
    """调用 analyze_polymarket_tracker.compute + format_ws，生成与 v5 一致的累计持仓表。"""
    import analyze_polymarket_tracker as apt
    from openpyxl.styles import Alignment

    name = TRACKER_STYLE_SHEET
    if raw_input.empty:
        pd.DataFrame({"说明": ["无已执行成交，无法生成累计持仓明细（与 analyze_polymarket_tracker 一致）。"]}).to_excel(
            writer, sheet_name=name, index=False
        )
        return
    proc, meta = apt.compute(
        raw_input.copy(),
        _tracker_compute_args(position_value_denominator=position_value_denominator),
    )
    proc = _align_tracker_subtotals_with_cycle_snapshot(proc, cycle_df)
    if "市场标题" in proc.columns:
        proc = proc.drop(columns=["市场标题"])
    proc.to_excel(writer, sheet_name=name, index=False)
    ws = writer.sheets[name]
    apt.format_ws(ws, meta.get("marker_col"))
    # 可读性：将“阶段”列按值居中。
    phase_col_idx = None
    for idx, cell in enumerate(ws[1], start=1):
        if str(cell.value or "").strip() == "阶段":
            phase_col_idx = idx
            break
    if phase_col_idx is not None and ws.max_row >= 2:
        for row in range(2, ws.max_row + 1):
            c = ws.cell(row=row, column=phase_col_idx)
            c.alignment = Alignment(horizontal="center", vertical="center")


def _finalize_xlsx_workbook(path: Path, *, skip_sheet_titles: frozenset[str] | None = None) -> None:
    """表头加粗并冻结首行；跳过多样式工作表（已由 format_ws 处理）。"""
    from openpyxl import load_workbook
    from openpyxl.styles import Font

    skip = skip_sheet_titles or frozenset()
    wb = load_workbook(path)
    for ws in wb.worksheets:
        if ws.title in skip:
            continue
        if ws.max_row == 0:
            continue
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
    wb.save(path)


def _write_performance_report_xlsx(
    xlsx_path: Path,
    *,
    shown_dir: Path,
    total_cycles: int,
    total_profit: float,
    avg_profit: float,
    win_rate: float,
    max_drawdown: float,
    total_fees: float,
    total_executed: int,
    best_cycle: pd.Series | None,
    worst_cycle: pd.Series | None,
    winner_df: pd.DataFrame,
    net_direction_df: pd.DataFrame,
    direction_df: pd.DataFrame,
    enriched_summary: pd.DataFrame,
    cycle_df: pd.DataFrame,
    decision_df: pd.DataFrame,
    tail_per_cycle_df: pd.DataFrame | None = None,
    tail_overview_rows: list[tuple[str, object]] | None = None,
    report_source: str = "",
    position_value_denominator: float = DEFAULT_POSITION_VALUE_DENOMINATOR,
    phase_end_seconds: tuple[float, float, float] = DEFAULT_PHASE_END_SECONDS,
) -> None:
    overview_rows: list[tuple[str, object]] = []
    if report_source:
        overview_rows.append(("数据来源", report_source))
    overview_rows += [
        ("回放目录", str(shown_dir.as_posix())),
        ("周期数", int(total_cycles)),
        ("总净利润", float(total_profit)),
        ("平均单周期净利润", float(avg_profit)),
        ("胜率", float(win_rate)),
        ("最大回撤", float(max_drawdown)),
        ("总手续费", float(total_fees)),
        ("总执行成交数", int(total_executed)),
    ]
    if best_cycle is not None:
        overview_rows.append(
            (
                "最佳周期",
                f"{best_cycle['cycle_slug']} | 净利润 {float(best_cycle['total_net_profit']):.6f}",
            )
        )
    if worst_cycle is not None:
        overview_rows.append(
            (
                "最差周期",
                f"{worst_cycle['cycle_slug']} | 净利润 {float(worst_cycle['total_net_profit']):.6f}",
            )
        )
    if not winner_df.empty:
        overview_rows.append(
            (
                "周期赢家分布",
                "，".join(f"{row['winner']}={int(row['count'])}" for _, row in winner_df.iterrows()),
            )
        )
    if not net_direction_df.empty:
        overview_rows.append(
            (
                "周期结束净方向分布",
                "，".join(f"{row['net_direction']}={int(row['count'])}" for _, row in net_direction_df.iterrows()),
            )
        )
    if not direction_df.empty:
        overview_rows.append(
            (
                "执行方向分布摘要",
                "，".join(
                    f"{row['selected_outcome']} {row['selected_action']}={int(row['trades'])}笔/{float(row['shares']):.6f}份"
                    for _, row in direction_df.iterrows()
                ),
            )
        )

    overview_df = pd.DataFrame(overview_rows, columns=["项目", "值"])
    if tail_overview_rows:
        overview_df = pd.concat(
            [overview_df, pd.DataFrame(tail_overview_rows, columns=["项目", "值"])],
            ignore_index=True,
        )

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    tracker_raw = _batch_replay_decisions_to_tracker_input(
        cycle_df,
        decision_df,
        phase_end_seconds=phase_end_seconds,
    )
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        overview_df.to_excel(writer, sheet_name="概览", index=False)
        enriched_summary.to_excel(writer, sheet_name="分周期累计盈亏", index=False)
        direction_df.to_excel(writer, sheet_name="执行方向分布", index=False)
        winner_df.to_excel(writer, sheet_name="周期赢家分布", index=False)
        net_direction_df.to_excel(writer, sheet_name="周期净方向分布", index=False)
        full_cycle_prices = _build_cycle_price_frame(decision_df, cycle_df, complete_only=True)
        _write_cycle_price_sheet(
            writer,
            sheet_name="完整周期价格图",
            cycle_price_df=full_cycle_prices,
            chart_title_prefix="完整周期价格折线图",
        )
        if tail_per_cycle_df is not None and not tail_per_cycle_df.empty:
            tail_per_cycle_df.to_excel(writer, sheet_name="尾盘窗口成交汇总", index=False)
        _append_tracker_style_sheet(
            writer,
            tracker_raw,
            cycle_df,
            position_value_denominator=position_value_denominator,
        )

    _finalize_xlsx_workbook(xlsx_path, skip_sheet_titles=frozenset({TRACKER_STYLE_SHEET}))


def _today_suffix() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _normalize_report_name_prefix(report_name_prefix: str) -> str:
    """归一化前缀，缩短常用名称并清理空白。"""
    safe_prefix = str(report_name_prefix or "").strip().replace(" ", "_")
    if not safe_prefix:
        return ""
    alias = {
        "simulation_streaming": "sim",
        "simulation": "sim",
    }
    return alias.get(safe_prefix, safe_prefix)


def build_performance_report_zh(
    resolved_batch_dir: Path,
    summary_df: pd.DataFrame,
    cycle_df: pd.DataFrame,
    decision_df: pd.DataFrame,
    output_path: Path,
    *,
    display_batch_dir: Path | None = None,
    cycle_range: str = "",
    cycle_count: int = 0,
    report_source: str = "",
    report_name_prefix: str = "",
    capital_reset_mode: str = "cumulative",
    starting_cash: float | None = None,
    position_value_denominator: float = DEFAULT_POSITION_VALUE_DENOMINATOR,
    phase_end_seconds: tuple[float, float, float] = DEFAULT_PHASE_END_SECONDS,
) -> Path:
    direction_df = _summarize_direction_distribution(decision_df)
    winner_df = _value_counts_frame(cycle_df.get("winner", pd.Series(dtype=object)), "winner")
    net_direction_df = _value_counts_frame(cycle_df.get("net_direction", pd.Series(dtype=object)), "net_direction")

    summary_df = summary_df.sort_values("cycle_slug").reset_index(drop=True)
    profits = [float(value) for value in summary_df.get("total_net_profit", pd.Series(dtype=float)).tolist()]
    total_cycles = len(summary_df)
    total_profit = float(sum(profits))
    avg_profit = float(summary_df["total_net_profit"].mean()) if total_cycles else 0.0
    win_rate = float((pd.to_numeric(summary_df["total_net_profit"], errors="coerce").fillna(0.0) > 0).mean()) if total_cycles else 0.0
    max_drawdown = _compute_max_drawdown(profits)
    total_fees = float(pd.to_numeric(summary_df.get("total_fees", 0.0), errors="coerce").fillna(0.0).sum()) if total_cycles else 0.0
    total_executed = int(pd.to_numeric(summary_df.get("executed_trades", 0), errors="coerce").fillna(0).sum()) if total_cycles else 0

    best_cycle = summary_df.sort_values("total_net_profit", ascending=False).iloc[0] if total_cycles else None
    worst_cycle = summary_df.sort_values("total_net_profit", ascending=True).iloc[0] if total_cycles else None

    enriched_summary = summary_df.copy()
    if total_cycles:
        enriched_summary["cumulative_profit"] = pd.to_numeric(enriched_summary["total_net_profit"], errors="coerce").fillna(0.0).cumsum()
        if {"account_cash", "total_net_profit"}.issubset(enriched_summary.columns):
            account_cash = pd.to_numeric(enriched_summary["account_cash"], errors="coerce")
            total_net_profit_col = pd.to_numeric(enriched_summary["total_net_profit"], errors="coerce").fillna(0.0)
            mode = str(capital_reset_mode or "").strip().lower()
            if mode == "fixed" and starting_cash is not None:
                implied_start_cash = pd.Series(float(starting_cash), index=enriched_summary.index, dtype=float)
            else:
                implied_start_cash = account_cash - total_net_profit_col
            enriched_summary["implied_start_cash"] = implied_start_cash
            first_start_cash = implied_start_cash.dropna().iloc[0] if implied_start_cash.notna().any() else 0.0
            enriched_summary["estimated_cash_from_cum"] = first_start_cash + enriched_summary["cumulative_profit"]

    suffix = _today_suffix()
    base_name = "batch_replay_performance_report"
    if report_name_prefix:
        safe_prefix = _normalize_report_name_prefix(report_name_prefix)
        if safe_prefix:
            base_name = f"{safe_prefix}_{base_name}"
    _vtag = get_version_tag()
    report_xlsx = output_path / f"{base_name}_{_vtag}_{suffix}.xlsx"
    shown_dir = display_batch_dir or resolved_batch_dir

    tail_per_cycle, tail_overview = summarize_tail_window_executions(
        decision_df, cycle_df, last_minute_seconds=30
    )

    _write_performance_report_xlsx(
        report_xlsx,
        shown_dir=shown_dir,
        total_cycles=total_cycles,
        total_profit=total_profit,
        avg_profit=avg_profit,
        win_rate=win_rate,
        max_drawdown=max_drawdown,
        total_fees=total_fees,
        total_executed=total_executed,
        best_cycle=best_cycle,
        worst_cycle=worst_cycle,
        winner_df=winner_df,
        net_direction_df=net_direction_df,
        direction_df=direction_df,
        enriched_summary=enriched_summary,
        cycle_df=cycle_df,
        decision_df=decision_df,
        tail_per_cycle_df=tail_per_cycle,
        tail_overview_rows=tail_overview,
        report_source=report_source,
        position_value_denominator=position_value_denominator,
        phase_end_seconds=phase_end_seconds,
    )

    return report_xlsx


def _find_latest_report(report_dir: Path) -> Path | None:
    """在目录中查找最新的绩效报告 xlsx 文件。"""
    candidates = sorted(
        report_dir.glob("*batch_replay_performance_report_*.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    return candidates[0] if candidates else None


def build_comparison_report_zh(
    sim_report_path: Path,
    live_report_path: Path,
    output_path: Path,
) -> Path:
    """生成「模拟下单测试 vs 实盘行情验证」对比报告。

    从两份绩效报告中提取概览数据和分周期数据，生成一份对比 Excel。
    """
    output_path.mkdir(parents=True, exist_ok=True)

    # --- 读取两份报告的概览 sheet ---
    sim_overview = pd.read_excel(sim_report_path, sheet_name="概览")
    live_overview = pd.read_excel(live_report_path, sheet_name="概览")

    # 构建并排对比行
    comparison_rows: list[dict[str, object]] = []
    sim_map: dict[str, object] = dict(zip(sim_overview["项目"], sim_overview["值"]))
    live_map: dict[str, object] = dict(zip(live_overview["项目"], live_overview["值"]))
    all_keys = list(dict.fromkeys(list(sim_map.keys()) + list(live_map.keys())))
    # 过滤掉纯路径 / 来源标识行
    skip_keys = {"回放目录", "数据来源"}
    for key in all_keys:
        if key in skip_keys:
            continue
        sim_val = sim_map.get(key, "—")
        live_val = live_map.get(key, "—")
        diff = ""
        try:
            sv = float(sim_val)  # type: ignore[arg-type]
            lv = float(live_val)  # type: ignore[arg-type]
            diff_val = lv - sv
            diff = f"{diff_val:+.6f}"
        except (ValueError, TypeError):
            pass
        comparison_rows.append({
            "指标": key,
            "模拟下单": sim_val,
            "实盘验证": live_val,
            "差异 (实盘-模拟)": diff,
        })

    comparison_df = pd.DataFrame(comparison_rows)

    # --- 相对真实性评分（以实盘验证报告为基准）---
    # 评分思路：模拟报告与实盘报告在关键指标上的偏差越小，分数越高；实盘侧固定为 100 作为基准。
    def _as_float(value: object) -> float | None:
        try:
            return float(value)  # type: ignore[arg-type]
        except (TypeError, ValueError):
            return None

    metric_specs: list[tuple[str, float, str]] = [
        ("总净利润", 0.35, "relative"),
        ("胜率", 0.20, "absolute"),
        ("最大回撤", 0.20, "relative"),
        ("总执行成交数", 0.15, "relative"),
        ("总手续费", 0.10, "relative"),
    ]
    weighted_penalty = 0.0
    weight_sum = 0.0
    metric_detail_rows: list[dict[str, object]] = []
    for key, weight, mode in metric_specs:
        sim_v = _as_float(sim_map.get(key))
        live_v = _as_float(live_map.get(key))
        if sim_v is None or live_v is None:
            metric_detail_rows.append(
                {
                    "指标": key,
                    "模拟值": sim_map.get(key, "—"),
                    "实盘值": live_map.get(key, "—"),
                    "偏差": "—",
                    "加权惩罚": "—",
                    "说明": "缺少可数值化字段，未纳入评分",
                }
            )
            continue
        if mode == "absolute":
            penalty = min(1.0, abs(sim_v - live_v))
            diff_text = f"{sim_v - live_v:+.6f}"
        else:
            denom = max(abs(live_v), 1.0)
            penalty = min(1.0, abs(sim_v - live_v) / denom)
            diff_text = f"{sim_v - live_v:+.6f} (相对 {abs(sim_v - live_v) / denom:.2%})"
        weighted_penalty += weight * penalty
        weight_sum += weight
        metric_detail_rows.append(
            {
                "指标": key,
                "模拟值": sim_v,
                "实盘值": live_v,
                "偏差": diff_text,
                "加权惩罚": round(weight * penalty, 6),
                "说明": f"权重={weight:.2f}，惩罚∈[0,1]",
            }
        )
    if weight_sum > 0:
        sim_auth_score = max(0.0, min(100.0, 100.0 * (1.0 - weighted_penalty / weight_sum)))
        coverage = weight_sum / sum(weight for _, weight, _ in metric_specs)
    else:
        sim_auth_score = 0.0
        coverage = 0.0
    authenticity_df = pd.DataFrame(
        [
            {
                "对象": "实盘验证报告",
                "真实性评分": 100.0,
                "指标覆盖率": "100%",
                "评分说明": "作为对比基准（reference）",
            },
            {
                "对象": "模拟下单报告",
                "真实性评分": round(sim_auth_score, 2),
                "指标覆盖率": f"{coverage:.0%}",
                "评分说明": "基于关键指标相对偏差的加权评分（越接近实盘越高）",
            },
        ]
    )
    metric_detail_df = pd.DataFrame(metric_detail_rows)

    # 将评分结果同步附加到「对比概览」末尾，便于一页查看核心结论。
    score_rows = pd.DataFrame(
        [
            {
                "指标": "真实性评分（实盘基准）",
                "模拟下单": "—",
                "实盘验证": 100.0,
                "差异 (实盘-模拟)": "基准",
            },
            {
                "指标": "真实性评分（模拟贴近度）",
                "模拟下单": round(sim_auth_score, 2),
                "实盘验证": 100.0,
                "差异 (实盘-模拟)": f"{100.0 - round(sim_auth_score, 2):+.2f}",
            },
        ]
    )
    comparison_df = pd.concat([comparison_df, score_rows], ignore_index=True)

    # --- 元信息 ---
    meta_rows = [
        ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("模拟报告", str(sim_report_path.name)),
        ("实盘报告", str(live_report_path.name)),
    ]
    meta_df = pd.DataFrame(meta_rows, columns=["项目", "值"])

    # --- 读取分周期数据 ---
    try:
        sim_cycles = pd.read_excel(sim_report_path, sheet_name="分周期累计盈亏")
    except Exception:
        sim_cycles = pd.DataFrame()
    try:
        live_cycles = pd.read_excel(live_report_path, sheet_name="分周期累计盈亏")
    except Exception:
        live_cycles = pd.DataFrame()

    # --- 写出 ---
    report_name = f"sim_vs_live_comparison_zh_{get_version_tag()}_{_today_suffix()}.xlsx"
    report_xlsx = output_path / report_name
    with pd.ExcelWriter(report_xlsx, engine="openpyxl") as writer:
        meta_df.to_excel(writer, sheet_name="报告信息", index=False)
        comparison_df.to_excel(writer, sheet_name="对比概览", index=False)
        authenticity_df.to_excel(writer, sheet_name="真实性评分", index=False)
        metric_detail_df.to_excel(writer, sheet_name="真实性评分明细", index=False)
        if not sim_cycles.empty:
            sim_cycles.to_excel(writer, sheet_name="模拟下单_分周期", index=False)
        if not live_cycles.empty:
            live_cycles.to_excel(writer, sheet_name="实盘验证_分周期", index=False)
    _finalize_xlsx_workbook(report_xlsx)
    return report_xlsx


def _write_batch_trade_process_xlsx(
    *,
    input_dir: Path,
    cycle_df: pd.DataFrame,
    decision_df: pd.DataFrame,
    output_path: Path,
) -> Path:
    """交易过程 Excel：元数据、周期快照、决策流水。"""
    xlsx_path = output_path / f"batch_replay_trade_process_zh_{get_version_tag()}_{_today_suffix()}.xlsx"
    meta_df = pd.DataFrame([("数据目录", input_dir.as_posix())], columns=["项", "值"])
    snap_df = cycle_df.copy() if not cycle_df.empty else pd.DataFrame()

    dec = decision_df.copy()
    if not dec.empty:
        if "timestamp" in dec.columns:
            dec["_ts"] = pd.to_datetime(dec["timestamp"], errors="coerce", utc=True)
            dec = dec.sort_values(["cycle_slug", "_ts"], kind="mergesort").drop(columns=["_ts"])
        elif "cycle_slug" in dec.columns:
            dec = dec.sort_values("cycle_slug", kind="mergesort")
        cols = [c for c in TRADE_PROCESS_PREF_COLS if c in dec.columns]
        if cols:
            if "cycle_slug" in dec.columns:
                order = ["cycle_slug"] + cols
                dec_out = dec[[c for c in order if c in dec.columns]]
            else:
                dec_out = dec[cols]
        else:
            dec_out = dec.copy()
    else:
        dec_out = pd.DataFrame(columns=["cycle_slug"])

    # Clarify "market direction" (event outcome) vs "strategy direction" (selected outcome).
    dec_out = dec_out.rename(
        columns={
            "market_outcome": "市场方向",
            "selected_outcome": "策略方向",
            "selected_action": "策略动作",
        }
    )

    if "timestamp" in dec_out.columns and "cycle_slug" in dec_out.columns and not dec_out.empty:
        ts_u = pd.to_datetime(dec_out["timestamp"], errors="coerce", utc=True)
        abs_open = dec_out["cycle_slug"].map(absolute_cycle_open_timestamp_utc_for_cycle_id)
        fb_min = dec_out.groupby("cycle_slug", dropna=False)["timestamp"].transform(
            lambda s: pd.to_datetime(s, errors="coerce", utc=True).min()
        )
        cycle_ref = abs_open.where(abs_open.notna(), fb_min)
        elapsed_series = pd.Series(
            [_format_elapsed(t, c) for t, c in zip(ts_u, cycle_ref)],
            index=dec_out.index,
            dtype=object,
        )
        pos = int(dec_out.columns.get_loc("cycle_slug")) + 1
        dec_out.insert(pos, _TRADE_PROCESS_ELAPSED_COL, elapsed_series)

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    cycle_prices = _build_cycle_price_frame(decision_df, cycle_df, complete_only=False)
    include_charts = _resolve_trade_process_chart_mode(cycle_df=cycle_df, decision_df=decision_df)
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        meta_df.to_excel(writer, sheet_name="元数据", index=False)
        snap_df.to_excel(writer, sheet_name="周期快照", index=False)
        dec_out.to_excel(writer, sheet_name="决策流水", index=False)
        _write_cycle_price_sheet(
            writer,
            sheet_name="周期价格与图表",
            cycle_price_df=cycle_prices,
            chart_title_prefix="回放周期价格折线图",
            include_charts=include_charts,
        )
        _apply_basic_sheet_style(writer)
    if not include_charts:
        print("  ℹ 交易过程图表已自动关闭（大批量优化模式）。可设置 POLYNET_TRADE_PROCESS_CHARTS=1 强制开启。")
    return xlsx_path


def write_batch_trade_process_zh(
    *,
    input_dir: Path,
    cycle_df: pd.DataFrame,
    decision_df: pd.DataFrame,
    output_path: Path,
) -> Path:
    """按周期汇总决策流水，生成 Excel（与绩效报告配套）。"""
    if "cycle_slug" not in cycle_df.columns and not cycle_df.empty:
        raise ValueError("cycle_df 需要包含 cycle_slug 列")
    if "cycle_slug" not in decision_df.columns and not decision_df.empty:
        raise ValueError("decision_df 需要包含 cycle_slug 列")

    return _write_batch_trade_process_xlsx(
        input_dir=input_dir,
        cycle_df=cycle_df,
        decision_df=decision_df,
        output_path=output_path,
    )


def _cleanup_batch_replay_markdown(output_path: Path) -> None:
    """移除历史或并发生成的 Markdown 产物，目录中只保留 Excel。"""
    for name in ("batch_replay_performance_report_zh.md", "batch_replay_trade_process_zh.md"):
        p = output_path / name
        try:
            if p.exists():
                p.unlink()
        except OSError:
            pass


def _dedupe_newest_xlsx_by_prefix(output_dir: Path, filename_prefix: str) -> int:
    """``{prefix}_*.xlsx`` 只保留修改时间最新的一份，返回删除文件数。"""
    paths = [p for p in output_dir.glob(f"{filename_prefix}_*.xlsx") if p.is_file()]
    if len(paths) <= 1:
        return 0
    paths.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    removed = 0
    for p in paths[1:]:
        try:
            p.unlink()
            removed += 1
        except OSError:
            pass
    return removed


def cleanup_polymarket_live_record_job_artifacts(
    cycle_record_dir: str | Path,
    *,
    batch_subdir: str = "batch_replay_outputs",
) -> None:
    """实盘验证 ``--record-events-dir`` 目录收尾：删 staging、合并重复的分析 Excel。

    - ``replay_new_cycles_*``：批量回放前临时复制的新周期树，与落盘的 ``btc-updown-5m-*`` 重复，可删。
    - ``batch_subdir`` 下同一前缀的带时间戳 xlsx 多次生成时，各前缀只保留最新一份。
    """
    root = Path(cycle_record_dir)
    if not root.is_dir():
        return
    for staging in sorted(root.glob("replay_new_cycles_*")):
        if staging.is_dir():
            shutil.rmtree(staging, ignore_errors=True)
    batch_dir = root / batch_subdir
    if not batch_dir.is_dir():
        return
    for prefix in (
        "real_batch_replay_performance_report",
        "simulation_batch_replay_performance_report",
        "sim_vs_live_comparison_zh",
        "batch_replay_trade_process_zh",
        "sim_batch_replay_performance_report",
    ):
        _dedupe_newest_xlsx_by_prefix(batch_dir, prefix)


def build_report(
    batch_dir: str | Path,
    output_dir: str | Path | None = None,
    *,
    config_path: str | Path = "configs/strategy.yaml",
    position_value_denominator: float | None = None,
) -> Path:
    resolved_batch_dir = _resolve_batch_replay_dir(batch_dir)
    cycle_dirs = _discover_cycle_result_dirs(resolved_batch_dir)
    if not cycle_dirs:
        raise RuntimeError(f"未在目录下找到任何周期回放结果: {resolved_batch_dir}")

    output_path = Path(output_dir) if output_dir else resolved_batch_dir
    output_path.mkdir(parents=True, exist_ok=True)

    summary_df = _load_summary(resolved_batch_dir, cycle_dirs)
    cycle_df, decision_df = _load_cycle_frames(cycle_dirs)
    write_batch_trade_process_zh(input_dir=resolved_batch_dir, cycle_df=cycle_df, decision_df=decision_df, output_path=output_path)
    report_path = build_performance_report_zh(
        resolved_batch_dir,
        summary_df,
        cycle_df,
        decision_df,
        output_path,
        display_batch_dir=resolved_batch_dir,
        position_value_denominator=resolve_position_value_denominator(
            config_path=config_path,
            explicit=position_value_denominator,
        ),
        phase_end_seconds=resolve_phase_end_seconds(config_path=config_path),
    )
    _cleanup_batch_replay_markdown(output_path)
    return report_path


def main() -> int:
    args = parse_args()
    report_path = build_report(
        args.input_dir,
        args.output_dir,
        config_path=args.config,
        position_value_denominator=args.position_value_denominator,
    )
    print(f"已生成中文总绩效报告: {report_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
