from __future__ import annotations

import argparse
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook


UP_HINTS = ("up", "yes", "涨", "看涨", "做多", "多")
DOWN_HINTS = ("down", "no", "跌", "看跌", "做空", "空")


def to_float(v: object) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    s = str(v).strip()
    if not s:
        return 0.0
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return 0.0


def f3(v: float) -> int | float:
    r = round(float(v), 3)
    if abs(r) < 1e-10:
        return 0
    if abs(r - round(r)) < 1e-10:
        return int(round(r))
    return r


def parse_outcome(v: object) -> Optional[str]:
    t = str(v or "").strip().lower()
    if not t:
        return None
    if any(h in t for h in UP_HINTS):
        return "up"
    if any(h in t for h in DOWN_HINTS):
        return "down"
    return None


def pick_first_col(header_to_col: dict[str, int], candidates: tuple[str, ...]) -> Optional[int]:
    for c in candidates:
        if c in header_to_col:
            return header_to_col[c]
    # also allow "normalized" matches
    norm = {str(k).strip().lower().replace(" ", ""): v for k, v in header_to_col.items()}
    for c in candidates:
        cn = str(c).strip().lower().replace(" ", "")
        if cn in norm:
            return norm[cn]
    return None


def update_one_sheet(ws, *, sheet_label: str) -> bool:
    # header row = 1
    header_to_col: dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        v = ws.cell(1, col_idx).value
        if v is None:
            continue
        header_to_col[str(v)] = col_idx

    if "Down已成交差价盈亏" not in header_to_col:
        return False
    if "浮动盈亏" in header_to_col:
        return False

    idx_down = header_to_col["Down已成交差价盈亏"]
    idx_new = idx_down + 1

    idx_price = header_to_col.get("成交价格")
    idx_outcome = header_to_col.get("结果代币类型")
    idx_cycle = header_to_col.get("时间周期")

    idx_marker = pick_first_col(
        header_to_col,
        (
            "下注时间距开盘差(分，秒)",
            "下注时间距开盘差（分，秒）",
            "下注时间距开盘差(分，秒）",
            "下注时间距开盘差",
        ),
    )

    idx_up_qty = header_to_col.get("Up积累份数")
    idx_dn_qty = header_to_col.get("Down积累份数")
    idx_up_avg = pick_first_col(header_to_col, ("Up的加权均价", "Up加权均价"))
    idx_dn_avg = pick_first_col(header_to_col, ("Down的加权均价", "Down加权均价"))

    needed = {
        "成交价格": idx_price,
        "结果代币类型": idx_outcome,
        "时间周期": idx_cycle,
        "下注时间距开盘差(分，秒)(或等价列)": idx_marker,
        "Up积累份数": idx_up_qty,
        "Up加权均价(或Up的加权均价)": idx_up_avg,
        "Down积累份数": idx_dn_qty,
        "Down加权均价(或Down的加权均价)": idx_dn_avg,
    }
    missing = [k for k, v in needed.items() if v is None]
    if missing:
        raise ValueError(f"[{sheet_label}] 缺少列，无法计算浮动盈亏: {missing}")

    # Insert new column after "Down已成交差价盈亏".
    ws.insert_cols(idx_new)
    ws.cell(1, idx_new).value = "浮动盈亏"
    # NOTE: 不强行复制 openpyxl 样式，避免 StyleProxy 在不同对象间赋值导致异常。

    last_float_by_cycle: dict[object, int | float] = {}

    def compute_for_row(r: int) -> int | float:
        price = to_float(ws.cell(r, idx_price).value)
        outcome = parse_outcome(ws.cell(r, idx_outcome).value)
        up_px: float
        dn_px: float
        if outcome == "up":
            up_px = price
            dn_px = 1.0 - price
        elif outcome == "down":
            dn_px = price
            up_px = 1.0 - price
        else:
            # 未识别方向：返回 0（一般用于 subtotal 行，后面会从 last_float_by_cycle 回填）
            return 0

        up_held = to_float(ws.cell(r, idx_up_qty).value)
        up_avg = to_float(ws.cell(r, idx_up_avg).value)
        dn_held = to_float(ws.cell(r, idx_dn_qty).value)
        dn_avg = to_float(ws.cell(r, idx_dn_avg).value)

        return f3(up_held * (up_px - up_avg) + dn_held * (dn_px - dn_avg))

    for r in range(2, ws.max_row + 1):
        cycle_val = ws.cell(r, idx_cycle).value
        marker_val = ws.cell(r, idx_marker).value if idx_marker else None
        marker_text = str(marker_val or "")
        is_subtotal = "【周期小计】" in marker_text

        if is_subtotal:
            if cycle_val in last_float_by_cycle:
                val = last_float_by_cycle[cycle_val]
            else:
                continue
        else:
            val = compute_for_row(r)
            last_float_by_cycle[cycle_val] = val

        c = ws.cell(r, idx_new, val)
        c.number_format = "0.000"

    return True


def update_xlsx(path: Path, *, sheets: Optional[list[str]] = None) -> None:
    wb = load_workbook(path)
    updated_any = False
    for name in wb.sheetnames:
        if sheets is not None and name not in sheets:
            continue
        ws = wb[name]
        updated = update_one_sheet(ws, sheet_label=name)
        updated_any = updated_any or updated
    if updated_any:
        wb.save(path)


def main() -> int:
    ap = argparse.ArgumentParser(description="为现有 Excel 补齐“浮动盈亏”列")
    ap.add_argument("--processed-v5", type=Path, required=False, default=Path("data/processed/polymarket_tracker_collection_with_accumulated_shares_v5.xlsx"))
    ap.add_argument("--performance-xlsx", type=Path, required=False, default=Path("artifacts/live/record_job/batch_replay_outputs/batch_replay_performance_report_zh.xlsx"))
    args = ap.parse_args()

    if args.processed_v5.exists():
        update_xlsx(args.processed_v5, sheets=None)
        print(f"updated: {args.processed_v5}")
    else:
        print(f"skip (not found): {args.processed_v5}")

    if args.performance_xlsx.exists():
        # only the tracker sheet
        update_xlsx(args.performance_xlsx, sheets=["分周期执行交易流水"])
        print(f"updated: {args.performance_xlsx} (sheet=分周期执行交易流水)")
    else:
        print(f"skip (not found): {args.performance_xlsx}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())

