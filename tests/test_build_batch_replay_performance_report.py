from __future__ import annotations

from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

from scripts.build_batch_replay_performance_report import (
    TRACKER_STYLE_SHEET,
    build_performance_report_zh,
    build_report,
    _compute_max_drawdown,
)


def test_compute_max_drawdown_from_cycle_profit_series() -> None:
    profits = [3.0, -5.0, 2.0, -1.0]

    value = _compute_max_drawdown(profits)

    assert value == 5.0


def test_build_report_writes_xlsx_and_trade_process(tmp_path) -> None:
    batch_dir = tmp_path / "batch_replay_outputs"
    cycle_dir = batch_dir / "btc-updown-5m-1774147200"
    cycle_dir.mkdir(parents=True)

    pd.DataFrame(
        [
            {
                "cycle_slug": "btc-updown-5m-1774147200",
                "executed_trades": 3,
                "accepted_signals": 3,
                "blocked_signals": 0,
                "total_net_profit": 1.5,
                "total_fees": 0.1,
                "winner": "up",
                "account_cash": 101.5,
            }
        ]
    ).to_csv(batch_dir / "batch_replay_summary.csv", index=False, encoding="utf-8-sig")

    pd.DataFrame(
        [
            {
                "market_id": "btc-up-or-down-5m",
                "cycle_id": "btc-updown-5m-1774147200",
                "net_direction": "Up",
                "winner": "up",
                "account_cash": 101.5,
            }
        ]
    ).to_csv(cycle_dir / "cycles.csv", index=False, encoding="utf-8-sig")

    pd.DataFrame(
        [
            {
                "timestamp": datetime(2026, 3, 20, 12, 0, 0),
                "selected_outcome": "up",
                "selected_action": "buy",
                "selected_shares": 10.0,
                "executed": True,
                "fill_price": 0.55,
                "fill_fee": 0.02,
            },
            {
                "timestamp": datetime(2026, 3, 20, 12, 0, 30),
                "selected_outcome": "up",
                "selected_action": "sell",
                "selected_shares": 5.0,
                "executed": True,
                "fill_price": 0.62,
                "fill_fee": 0.01,
            },
            {
                "timestamp": datetime(2026, 3, 20, 12, 4, 35),
                "selected_outcome": "down",
                "selected_action": "buy",
                "selected_shares": 2.0,
                "executed": True,
                "fill_price": 0.40,
                "fill_fee": 0.001,
            },
        ]
    ).to_csv(cycle_dir / "decisions.csv", index=False, encoding="utf-8-sig")

    pd.DataFrame(
        [
            {
                "total_net_profit": 1.5,
                "total_fees": 0.1,
                "executed_trades": 3,
                "accepted_signals": 3,
                "blocked_signals": 0,
            }
        ]
    ).to_csv(cycle_dir / "metrics.csv", index=False, encoding="utf-8-sig")

    report_path = build_report(batch_dir)

    assert report_path.suffix == ".xlsx"
    assert report_path.exists()
    xl = pd.ExcelFile(report_path)
    assert "概览" in xl.sheet_names
    overview = pd.read_excel(report_path, sheet_name="概览")
    lookup = {str(k): v for k, v in zip(overview["项目"], overview["值"])}
    assert int(float(lookup["周期数"])) == 1
    assert abs(float(lookup["总净利润"]) - 1.5) < 1e-9
    assert abs(float(lookup["胜率"]) - 1.0) < 1e-9
    assert abs(float(lookup["最大回撤"]) - 0.0) < 1e-9
    assert int(float(lookup["尾盘已执行成交笔数(合计)"])) == 1
    assert abs(float(lookup["尾盘手续费合计"]) - 0.001) < 1e-9
    assert abs(float(lookup["尾盘成交现金流净额(不含结算)"]) - (-(2.0 * 0.40 + 0.001))) < 1e-9

    assert "尾盘窗口成交汇总" in xl.sheet_names
    tail_df = pd.read_excel(report_path, sheet_name="尾盘窗口成交汇总")
    assert len(tail_df) == 1
    assert int(tail_df["tail_executed_trades"].iloc[0]) == 1

    assert TRACKER_STYLE_SHEET in xl.sheet_names
    tracker = pd.read_excel(report_path, sheet_name=TRACKER_STYLE_SHEET)
    cols = list(tracker.columns)
    assert "成交价格" in cols
    assert "同向成交价波动幅度(%)" in cols
    assert cols.index("同向成交价波动幅度(%)") == cols.index("成交价格") + 1
    assert "决策原因" in cols
    assert "Up积累份数" in tracker.columns
    assert "周期净利润" in tracker.columns
    ratio_series = pd.to_numeric(tracker["持仓价值占比"], errors="coerce").dropna()
    assert not ratio_series.empty
    # 导出链路会在写回前统一 round(3)，断言按三位小数口径比较。
    expected_ratio = round((10.0 * 0.55) / 85.0, 3)
    assert abs(float(ratio_series.iloc[0]) - expected_ratio) < 1e-9
    marker_col = "下注时间距开盘差(分，秒)"
    assert marker_col in tracker.columns
    sub = tracker[tracker[marker_col].astype(str) == "【周期小计】"]
    assert not sub.empty
    assert pd.notna(sub["未平仓UP盈亏"].iloc[0])
    assert pd.notna(sub["未平仓Down盈亏"].iloc[0])
    assert pd.notna(sub["周期净利润"].iloc[0])
    assert "完整周期价格图" in xl.sheet_names

    wb = load_workbook(report_path)
    assert len(wb["完整周期价格图"]._charts) == 1

    assert not (batch_dir / "batch_replay_performance_report_zh.md").exists()
    assert not (batch_dir / "batch_replay_trade_process_zh.md").exists()
    trade_files = list(batch_dir.glob("batch_replay_trade_process_zh_*.xlsx"))
    assert trade_files
    txl = pd.ExcelFile(trade_files[0])
    assert "元数据" in txl.sheet_names
    assert "周期快照" in txl.sheet_names
    assert "决策流水" in txl.sheet_names
    assert "周期价格与图表" in txl.sheet_names

    twb = load_workbook(trade_files[0])
    assert len(twb["周期价格与图表"]._charts) == 1
    assert not (batch_dir / "batch_replay_direction_distribution.csv").exists()
    assert not (batch_dir / "batch_replay_summary_enriched.csv").exists()


def test_build_performance_report_fixed_mode_implied_start_cash(tmp_path) -> None:
    output_dir = tmp_path / "report_out"
    output_dir.mkdir(parents=True)

    summary_df = pd.DataFrame(
        [
            {"cycle_slug": "c1", "total_net_profit": -28.232, "total_fees": 0.0, "executed_trades": 2, "account_cash": 71.768},
            {"cycle_slug": "c2", "total_net_profit": 9.572, "total_fees": 0.0, "executed_trades": 2, "account_cash": 81.34},
            {"cycle_slug": "c3", "total_net_profit": 10.428, "total_fees": 0.0, "executed_trades": 2, "account_cash": 91.768},
        ]
    )
    cycle_df = pd.DataFrame(
        [
            {"market_id": "btc-up-or-down-5m", "cycle_id": "c1", "winner": "down", "net_direction": "down"},
            {"market_id": "btc-up-or-down-5m", "cycle_id": "c2", "winner": "up", "net_direction": "up"},
            {"market_id": "btc-up-or-down-5m", "cycle_id": "c3", "winner": "up", "net_direction": "up"},
        ]
    )
    decision_df = pd.DataFrame(
        [
            {"market_id": "btc-up-or-down-5m", "cycle_id": "c1", "cycle_slug": "c1", "timestamp": datetime(2026, 3, 27, 10, 0, 0), "selected_action": "buy", "selected_outcome": "up", "selected_shares": 1.0, "executed": True, "fill_price": 0.5, "fill_fee": 0.0},
            {"market_id": "btc-up-or-down-5m", "cycle_id": "c2", "cycle_slug": "c2", "timestamp": datetime(2026, 3, 27, 10, 5, 0), "selected_action": "buy", "selected_outcome": "up", "selected_shares": 1.0, "executed": True, "fill_price": 0.5, "fill_fee": 0.0},
            {"market_id": "btc-up-or-down-5m", "cycle_id": "c3", "cycle_slug": "c3", "timestamp": datetime(2026, 3, 27, 10, 10, 0), "selected_action": "buy", "selected_outcome": "up", "selected_shares": 1.0, "executed": True, "fill_price": 0.5, "fill_fee": 0.0},
        ]
    )

    report_path = build_performance_report_zh(
        resolved_batch_dir=output_dir,
        summary_df=summary_df,
        cycle_df=cycle_df,
        decision_df=decision_df,
        output_path=output_dir,
        display_batch_dir=output_dir,
        capital_reset_mode="fixed",
        starting_cash=100.0,
    )

    xl = pd.ExcelFile(report_path)
    enriched = None
    for sheet_name in xl.sheet_names:
        candidate = pd.read_excel(report_path, sheet_name=sheet_name)
        if {"implied_start_cash", "estimated_cash_from_cum"}.issubset(set(candidate.columns)):
            enriched = candidate
            break
    assert enriched is not None
    implied = pd.to_numeric(enriched["implied_start_cash"], errors="coerce").dropna().tolist()
    assert implied == [100.0, 100.0, 100.0]
