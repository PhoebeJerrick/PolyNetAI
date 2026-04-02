from __future__ import annotations

from datetime import datetime, timedelta

from openpyxl import load_workbook

from polynet_ai.domain.models import FillEvent, TradeEvent
from polynet_ai.engine.live import LivePaperRunner, export_live_result
from polynet_ai.engine.replay import ReplayEngine
from polynet_ai.reporting.excel_export import get_version_tag
from polynet_ai.strategy.spec import StrategyConfig


def build_config() -> StrategyConfig:
    return StrategyConfig(
        raw={
            "cycle": {"cycle_seconds": 300, "last_minute_seconds": 60},
            "order_sizing": {"base_order_size": 8.0, "min_order_size": 2.0, "max_order_size": 60.0, "volatility_order_scale": 20.0},
            "exposure": {"max_abs_exposure_value": 200.0, "hedge_trigger_value": 50.0, "hedge_scale": 0.15, "max_strategy_trades_per_cycle": 12},
            "trend": {"min_trend_strength": 0.35, "trend_price_edge": 0.03, "trend_scale": 0.15},
            "grid": {"grid_low_percentile": 0.25, "grid_high_percentile": 0.75},
            "mean_reversion": {"up_buy_deviation": 0.10, "down_buy_deviation": 0.10, "mean_reversion_sell_up_deviation": 0.20, "mean_reversion_sell_down_deviation": 0.20, "deviation_scale": 45.0},
            "profit_taking": {"take_profit_up_deviation": 0.20, "take_profit_down_deviation": 0.20, "take_profit_fraction": 0.35},
            "stop_loss": {"stop_loss_cycle_loss": 20.0, "stop_loss_fraction": 0.50},
            "last_minute": {"last_minute_min_confidence": 0.60, "tail_profit_scale": 0.35, "tail_volatility_scale": 25.0, "max_tail_exposure": 40.0},
            "execution": {"fee_rate": 0.002, "slippage_bps": 10},
            "priorities": {"risk": 10, "last_minute": 20, "stop_loss": 30, "hedge": 40, "take_profit": 50, "grid": 60, "mean_reversion": 70, "trend": 80},
        }
    )


def test_live_runner_generates_snapshots_and_cycle_output() -> None:
    t0 = datetime(2026, 3, 20, 12, 0, 0)
    events = [
        TradeEvent("BTC", "cycle-a", t0, price=0.45, shares=10, outcome="up", action="buy"),
        TradeEvent("BTC", "cycle-a", t0 + timedelta(seconds=30), price=0.50, shares=8, outcome="up", action="buy"),
        TradeEvent("BTC", "cycle-a", t0 + timedelta(seconds=60), price=0.58, shares=7, outcome="up", action="buy"),
        TradeEvent("BTC", "cycle-b", t0 + timedelta(seconds=360), price=0.40, shares=6, outcome="down", action="buy"),
    ]
    sleeps: list[float] = []
    runner = LivePaperRunner(ReplayEngine(build_config()))
    result = runner.run(
        events,
        pace_factor=0.0,
        max_sleep_seconds=0.0001,
        status_every=0,
        sleep_fn=sleeps.append,
    )
    assert len(result.snapshot_df) == len(events)
    assert len(result.replay_result.cycle_df) == 2
    assert "account_cash" in result.snapshot_df.columns
    assert "up_last_price" in result.snapshot_df.columns
    assert "down_last_price" in result.snapshot_df.columns
    assert all(delay <= 0.01 for delay in sleeps)


def test_live_runner_stream_accepts_iterators() -> None:
    t0 = datetime(2026, 3, 20, 12, 0, 0)
    events = iter(
        [
            TradeEvent("btc-up-or-down-5m", "cycle-a", t0, price=0.45, shares=10, outcome="up", action="buy"),
            TradeEvent("btc-up-or-down-5m", "cycle-a", t0 + timedelta(seconds=45), price=0.49, shares=6, outcome="up", action="buy"),
            TradeEvent("btc-up-or-down-5m", "cycle-b", t0 + timedelta(seconds=320), price=0.52, shares=7, outcome="down", action="sell"),
        ]
    )
    runner = LivePaperRunner(ReplayEngine(build_config()))
    result = runner.run_stream(events, status_every=0)

    assert len(result.snapshot_df) == 3
    assert len(result.replay_result.cycle_df) == 2
    assert result.replay_result.cycle_df.iloc[0]["cycle_id"] == "cycle-a"
    assert result.replay_result.cycle_df.iloc[1]["cycle_id"] == "cycle-b"


def test_live_runner_stream_emits_events_in_consumed_order() -> None:
    t0 = datetime(2026, 3, 20, 12, 0, 0)
    events = iter(
        [
            TradeEvent("btc-up-or-down-5m", "cycle-a", t0, price=0.45, shares=10, outcome="up", action="buy"),
            TradeEvent("btc-up-or-down-5m", "cycle-a", t0 + timedelta(seconds=1), price=0.46, shares=6, outcome="down", action="sell"),
        ]
    )
    seen: list[tuple[str, float, str, str]] = []
    runner = LivePaperRunner(ReplayEngine(build_config()))

    runner.run_stream(
        events,
        status_every=0,
        on_event=lambda event: seen.append((event.cycle_id, event.price, event.outcome, event.action)),
    )

    assert seen == [
        ("cycle-a", 0.45, "up", "buy"),
        ("cycle-a", 0.46, "down", "sell"),
    ]


def test_export_live_result_writes_trade_ledger_excel(tmp_path) -> None:
    engine = ReplayEngine(build_config())
    t0 = datetime(2026, 3, 20, 12, 0, 0)
    engine.state_engine.apply_market_trade(
        TradeEvent("BTC", "cycle-a", t0, price=0.45, shares=10, outcome="up", action="buy")
    )
    engine.state_engine.apply_market_trade(
        TradeEvent("BTC", "cycle-a", t0 + timedelta(seconds=5), price=0.55, shares=8, outcome="down", action="buy")
    )
    engine.state_engine.apply_strategy_fill(
        FillEvent("BTC", "cycle-a", t0 + timedelta(seconds=10), price=0.46, shares=5, outcome="up", action="buy")
    )
    snapshot = engine.state_engine.snapshot()
    result = export_live_result(
        result=LivePaperRunner(engine)._build_live_result(
            cycle_rows=[],
            decision_rows=[
                {
                    "event_index": 1,
                    "market_id": "BTC",
                    "cycle_id": "cycle-a",
                    "timestamp": t0 + timedelta(seconds=10),
                    "selected_rule": "grid",
                    "selected_action": "buy",
                    "selected_outcome": "up",
                    "selected_shares": 5.0,
                    "risk_status": "accepted",
                    "risk_reason": "",
                    "executed": True,
                    "fill_price": 0.46,
                    "fill_fee": 0.0,
                    "cycle_net_profit": snapshot.cycle_net_profit,
                    "account_cash": engine.account.cash,
                }
            ],
            snapshot_rows=[
                {
                    "event_index": 1,
                    "timestamp": snapshot.timestamp,
                    "market_id": snapshot.market_id,
                    "cycle_id": snapshot.cycle_id,
                    "price": snapshot.price,
                    "up_balance": snapshot.up_balance,
                    "down_balance": snapshot.down_balance,
                    "total_position": snapshot.total_position,
                    "net_position": snapshot.net_position,
                    "net_direction": snapshot.net_direction,
                    "net_position_value": snapshot.net_position_value,
                    "up_avg_price": snapshot.up_avg_price,
                    "down_avg_price": snapshot.down_avg_price,
                    "up_realized_pnl": snapshot.up_realized_pnl,
                    "down_realized_pnl": snapshot.down_realized_pnl,
                    "unrealized_up_pnl": snapshot.unrealized_up_pnl,
                    "unrealized_down_pnl": snapshot.unrealized_down_pnl,
                    "cycle_net_profit": snapshot.cycle_net_profit,
                    "high_price": snapshot.high_price,
                    "low_price": snapshot.low_price,
                    "opening_price": snapshot.opening_price,
                    "up_last_price": snapshot.up_last_price,
                    "down_last_price": snapshot.down_last_price,
                    "market_trades": snapshot.market_trades,
                    "strategy_trades": snapshot.strategy_trades,
                    "account_cash": engine.account.cash,
                }
            ],
        ),
        output_dir=tmp_path,
    )
    workbook = load_workbook(result / f"trade_ledger_{get_version_tag()}.xlsx", read_only=True)
    assert workbook.sheetnames == ["BTC"]
    rows = list(workbook["BTC"].iter_rows(values_only=True))
    assert rows[0][0] == "下注时间距开盘差(分，秒)"
    assert rows[1][1] == "BTC"
    assert rows[1][3] == "up"
